"""
Excel ↔ PDF Reconciliation — Human-in-the-Loop, fully local.
The PDF (SMIR) is the source of truth. The Excel (GSIS-P) is derived from it and
may contain OCR errors. Integrates:
  • Phase 1: table-aware header metadata (Vendor / Part / Model / Issue Date)
  • Phase 2: context-aware OCR correction + merged-cell dedup table extraction
  • Phase 3: matching & reconciliation (fuzzy item match, numeric tolerance
    cross-checks, MIC category validation + auto-fill)
Run:  streamlit run reconcile.py
"""
from __future__ import annotations
import io, os, re, base64, hashlib, difflib, datetime, json, time, logging
import importlib.util as _modspec
import threading
from dataclasses import dataclass, field
from typing import Dict, List, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import uuid


os.environ.setdefault("HF_HUB_DISABLE_SYMLINKS", "1")
os.environ.setdefault("HF_HUB_DISABLE_SYMLINKS_WARNING", "1")

import streamlit as st

def generate_unique_id():
    """Generate a unique ID for each reconciliation session."""
    # Format: REC-YYYYMMDD-XXXX (where XXXX is random)
    today = datetime.datetime.now().strftime("%Y%m%d")
    random_part = str(uuid.uuid4())[:8].upper()
    return f"REC-{today}-{random_part}"

# ═══ FAST START: paint UI FIRST, load heavy engines AFTER ═══
st.set_page_config(page_title="Excel ↔ PDF Reconciliation", page_icon="🔍", layout="wide")
st.markdown("""<style>
.block-container {padding-top: 1.1rem; padding-bottom: 2rem; max-width: 1560px;}
.app-hero {background: linear-gradient(100deg,#1e3a8a 0%,#3b0764 100%);
color:#fff; padding:18px 26px; border-radius:16px; margin-bottom:14px;
box-shadow:0 6px 24px -8px rgba(59,7,100,.55);}
.app-hero h1 {margin:0; font-size:1.55rem; font-weight:750; letter-spacing:.2px;}
.app-hero p {margin:.3rem 0 0; opacity:.9; font-size:.9rem;}
div[data-testid="stDataFrame"] div, .stDataEditor {font-size:.9rem;}
.stButton>button {border-radius:10px; font-weight:600;}
.stButton>button:hover {border-color:#3b82f6;}
</style>""", unsafe_allow_html=True)
st.markdown("""<div class="app-hero">
<h1>🔍 Excel ↔ PDF Reconciliation</h1>
<p>The PDF is the source of truth · fully local · the 19-column schema is preserved</p>
</div>""", unsafe_allow_html=True)
_boot = st.empty()
_boot.info("⏳ Starting engines… (one-time load, a few seconds)")

def _has(m: str) -> bool:
    """Detect a library WITHOUT importing it (instant, no blank screen)."""
    try:
        return _modspec.find_spec(m) is not None
    except Exception:
        return False

# Core data libs (needed immediately, fast)
import pandas as pd
import numpy as np
import pdfplumber
from openpyxl import load_workbook
import fitz  # PyMuPDF

# Optional heavy libs: flags only — nothing heavy is imported here
RAPIDFUZZ_AVAILABLE = _has("rapidfuzz")
CV2_AVAILABLE       = _has("cv2")
TESSERACT_AVAILABLE = _has("pytesseract")
EASYOCR_AVAILABLE   = _has("easyocr")
CAMELOT_AVAILABLE   = _has("camelot")
TABULA_AVAILABLE    = _has("tabula")
FAISS_AVAILABLE     = _has("faiss")

fuzz = None
if RAPIDFUZZ_AVAILABLE:
    from rapidfuzz import fuzz

cv2 = None
if CV2_AVAILABLE:
    import cv2   # OpenCV now loads AFTER first paint → screen never blank

_boot.empty()

# Lazy loaders: heavy OCR engines load ONLY on first real use, then cached
@st.cache_resource
def _tesseract_mod():
    import pytesseract
    return pytesseract

@st.cache_resource
def _easyocr_reader():
    import easyocr
    return easyocr.Reader(["en"], gpu=False)

@st.cache_resource
def _camelot_mod():
    import camelot
    return camelot

@st.cache_resource
def _tabula_mod():
    import tabula
    return tabula

@st.cache_resource
def _faiss_mod():
    import faiss
    return faiss

def generate_unique_id():
    """Generate a unique ID for each reconciliation session."""
    # Format: REC-YYYYMMDD-XXXX (where XXXX is random)
    today = datetime.datetime.now().strftime("%Y%m%d")
    random_part = str(uuid.uuid4())[:8].upper()
    return f"REC-{today}-{random_part}"

# ════ 1. CONFIG & CONSTANTS ════
SCHEMA = ["VENDOR CODE","Part number","Model No.","Operation number","MIC Name",
 "Inspection Item","Parameter","Lower Limit","Target Value","Upper Limit",
 "Decimal Places","UOM","Inspection Method","Inspection Tool",
 "Info Field 1","Info Field 2","Info Field 3","Issue date","Long Text"]
NUMERIC_COLS = {"Operation number","Lower Limit","Target Value","Upper Limit","Decimal Places"}
REQUIRED_COLS = ["VENDOR CODE","Part number","Model No.","MIC Name","Inspection Item",
 "Parameter","Inspection Method","Inspection Tool","Info Field 1","Info Field 2",
 "Issue date","Long Text","Operation number"]
ALLOWED_EMPTY_COLS = ["Lower Limit","Target Value","Upper Limit","Decimal Places","UOM","Info Field 3"]
PDF_COLUMN_MAPPING = [("VENDOR CODE","📄 PDF Vendor"),("Part number","📄 PDF Part No"),
 ("Model No.","📄 PDF Model"),("Operation number","📄 PDF Serial No"),("MIC Name","📄 PDF MIC"),
 ("Inspection Item","📄 PDF Item"),("Parameter","📄 PDF Criteria"),
 ("Inspection Method","📄 PDF Method"),("Inspection Tool","📄 PDF Tool"),
 ("Info Field 1","📄 PDF Sampling"),("Info Field 2","📄 PDF Stage"),
 ("Info Field 3","📄 PDF Remarks"),("Issue date","📄 PDF Issue Date")]
MIC_PATTERN = ["Appearance","Dimension","Material","Performance","Revalidation"]
_PDF_READONLY_COLS = tuple(p for _, p in PDF_COLUMN_MAPPING)
_RESIDUAL_RE = re.compile(r":\s*(?:un)?select(?:ed)?\s*:", re.I)

# ════ 2. CONFIDENCE SCORING ════
def has_ocr_artifacts(text):
    return bool(text) and any(a in text for a in ["¥","¢","£","©","®","™","•","·","…","“","”","‘","’"])

def calculate_cell_confidence(value, column, pdf_value=None):
    if not value or str(value).strip() == "":
        return {"score": 0, "reasons": ["Empty cell"], "needs_review": True}
    s = str(value).strip(); score = 100; reasons = []
    if has_ocr_artifacts(s): score -= 25; reasons.append("OCR artifacts")
    if column == "Parameter":
        if parse_spec(s) is None:
            if not any(k in s.lower() for k in ("as per","sop","tis","drawing","mtg-","mtg/")):
                score -= 35; reasons.append("Invalid tolerance format")
        else:
            lim = parse_limits(s)
            if lim and lim.get("lower") is not None and lim.get("upper") is not None and lim["upper"]-lim["lower"] > 1000:
                score -= 10; reasons.append("Tolerance range too large")
    elif column == "MIC Name" and s not in MIC_PATTERN:
        score -= 30; reasons.append("Unknown MIC category")
    elif column == "Issue date" and not re.match(r'^\d{2}-\d{2}-\d{4}$', s):
        score -= 30; reasons.append("Invalid date format")
    elif column in ("Lower Limit","Upper Limit") and safe_float(s) is None:
        score -= 40; reasons.append("Invalid numeric value")
    elif column == "Decimal Places":
        try:
            f = float(s)
            if f != int(f): score -= 30; reasons.append("Must be integer")
        except (ValueError, TypeError):
            score -= 40; reasons.append("Must be numeric")
    if pdf_value and str(pdf_value).strip() and _norm(s) != _norm(pdf_value):
        if column == "Parameter" and _magnums(s) == _magnums(pdf_value):
            score -= 10; reasons.append("Same numbers, format differs from PDF")
        else:
            score -= 20; reasons.append("Differs from PDF")
    score = max(0, min(100, score))
    return {"score": score, "reasons": reasons, "needs_review": score < 70}

def get_row_confidence_summary(conf):
    if not conf: return {"avg": 0, "min": 0, "needs_review": True}
    sc = [c["score"] for c in conf.values()]
    return {"avg": int(sum(sc)/len(sc)), "min": min(sc),
            "needs_review": any(c["needs_review"] for c in conf.values())}

# ════ 3. HELPERS ════
def _norm(x): return re.sub(r"[^a-z0-9]", "", str(x).lower())
def safe_float(v):
    try: f = float(str(v).strip())
    except (TypeError, ValueError): return None
    return f if (f == f and f not in (float("inf"), float("-inf"))) else None
def _magnums(x): return sorted(re.findall(r"\d+(?:\.\d+)?", str(x)))
def strip_residuals(v):
    if v is None: return v
    s = str(v)
    if ":" not in s or not _RESIDUAL_RE.search(s): return v
    return re.sub(r"\s{2,}", " ", _RESIDUAL_RE.sub("", s)).strip()
def pdf_num_pages(b):
    try:
        d = fitz.open(stream=b, filetype="pdf"); n = len(d); d.close(); return n
    except Exception: return 0
def _parse_pages(spec, n):
    spec = (spec or "").strip()
    if not spec: return list(range(n))
    idx = set()
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            a, _, b = part.partition("-")
            if a.strip().isdigit() and b.strip().isdigit():
                for p in range(int(a), int(b)+1):
                    if 1 <= p <= n: idx.add(p-1)
        elif part.isdigit():
            p = int(part)
            if 1 <= p <= n: idx.add(p-1)
    return sorted(idx) if idx else list(range(n))
def subset_pdf(b, spec):
    try:
        d = fitz.open(stream=b, filetype="pdf"); n = len(d); idx = _parse_pages(spec, n)
        if not idx or len(idx) == n: d.close(); return b
        out = fitz.open()
        for i in idx: out.insert_pdf(d, from_page=i, to_page=i)
        r = out.tobytes(); d.close(); out.close(); return r
    except Exception: return b
@st.cache_data(show_spinner=False)
def pdf_page_pngs(b, dpi=200):
    out = []
    try:
        doc = fitz.open(stream=b, filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72), colorspace=fitz.csRGB)
            out.append(pix.tobytes("png"))
        doc.close()
    except Exception: pass
    return out
def _find_header_row(ws, columns):
    wanted = {_norm(c) for c in columns}; best, bh = 1, -1
    for r in range(1, min(ws.max_row, 15)+1):
        hits = sum(1 for cell in ws[r] if cell.value is not None and _norm(cell.value) in wanted)
        if hits > bh: best, bh = r, hits
    return best

def _pyval(v):
    """Convert numpy/pandas scalars to native Python types openpyxl can write."""
    if v is None: return None
    if hasattr(v, "item"):                      # numpy scalar -> python
        try: v = v.item()
        except Exception: pass
    try:
        if pd.isna(v): return None
    except Exception: pass
    s = str(v).strip()
    if s.lower() in ("", "nan", "none"): return None
    return v

def build_workbook(excel_bytes, work):
    # Path 1: try to preserve the original template
    try:
        wb = load_workbook(io.BytesIO(excel_bytes)); ws = wb.active
        hr = _find_header_row(ws, SCHEMA)
        col_at = {_norm(c.value): c.column for c in ws[hr] if c.value is not None}
        for i in range(len(work)):
            for name in SCHEMA:
                col = col_at.get(_norm(name))
                if not col: continue
                val = _pyval(work.iloc[i][name])
                if name == "Decimal Places" and safe_float(val) == 0: val = None
                elif name in NUMERIC_COLS and safe_float(val) is not None:
                    fv = safe_float(val)
                    val = int(fv) if name in ("Operation number","Decimal Places") and fv == int(fv) else float(fv)
                ws.cell(row=hr+1+i, column=col, value=val)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
    except Exception:
        # Path 2: template can't be re-saved -> build a clean 19-col workbook (guaranteed)
        from openpyxl import Workbook
        wb2 = Workbook(); ws2 = wb2.active; ws2.title = "GSIS-P"
        ws2.append(SCHEMA)
        for i in range(len(work)):
            row = []
            for name in SCHEMA:
                val = _pyval(work.iloc[i][name])
                if name in NUMERIC_COLS and safe_float(val) is not None:
                    fv = safe_float(val); val = int(fv) if fv == int(fv) else float(fv)
                row.append(val)
            ws2.append(row)
        buf = io.BytesIO(); wb2.save(buf); return buf.getvalue()

def _to_xlsx(df):
    df = df.copy()
    for c in df.columns: df[c] = df[c].map(_pyval)
    buf = io.BytesIO(); df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()

# ════ 4. PDF EXTRACTION (Phase 2) ════
_CATEGORIES = {"appearance","dimensionsinmm","dimensions","materialperformance",
 "materialperformanceasperdrawing","performanceofhose","performanceofhoseasperdrawing",
 "revalidation","epdmspongeprotector","elvcompliance","layout","mandrelshapeverification","netprotector"}
_MIC_KEYWORDS = {"appearance":"Appearance","dimension":"Dimension","performance":"Performance",
 "material":"Material","revalidation":"Revalidation","layout":"Revalidation"}
def _section_cat(text):
    n = _norm(text); n = re.sub(r"^(?:[ivxl]{1,4}|\d{1,3})(?=[a-z])", "", n)
    for kw, canon in _MIC_KEYWORDS.items():
        if n.startswith(kw): return canon
    return ""
_SKIP = ("inspectedby","verifiedby","approvedby","approvedbysmg","qaf5014","note","totalno",
 "lotquantity","finaljudgement","summaryof","vendorname","sno","inspectionitemspecification",
 "msil","modifiedby","revisionnumber","formno","documentclassification","checked","prepared",
 "descriptionmodified","inspectionitemcriteria")
def _digital_words_per_page(b):
    try:
        with pdfplumber.open(io.BytesIO(b)) as pdf:
            n = len(pdf.pages) or 1
            return sum(len((p.extract_text() or "").split()) for p in pdf.pages) / n
    except Exception: return 0.0

# ── Phase 2: context-aware OCR correction ──
_OCR_SYMBOL_FIX = {"•":".","·":".","…":"...","“":'"',"”":'"',"‘":"'","’":"'",
 "©":"(c)","®":"(R)","™":"TM","¢":"c","£":"L"}
_OCR_DIGIT_TO_LETTER = {"0":"O","1":"I","2":"Z","4":"A","5":"S","6":"G","8":"B","9":"g"}
_UNIT_SUFFIX = {"hr","hrs","hour","hours","lot","part","parts","month","week","day","mm","kg",
 "min","max","nos","no","pcs","shift","setup","yr","sec","lpm","hrc","hrb","digit","digits"}
def _fix_word_token(tok):
    alpha = sum(1 for ch in tok if ch.isalpha())
    digit = sum(1 for ch in tok if ch.isdigit())
    if not (alpha >= 3 and 1 <= digit <= 2 and alpha > digit): return tok
    m = re.match(r"^(\d{1,3})([A-Za-z]+)$", tok)
    if m and m.group(2).lower() in _UNIT_SUFFIX: return tok
    return "".join(_OCR_DIGIT_TO_LETTER.get(ch, ch) for ch in tok)
def post_process_ocr(text):
    if not text: return text
    for w, c in _OCR_SYMBOL_FIX.items(): text = text.replace(w, c)
    text = re.sub(r"[A-Za-z0-9]+", lambda m: _fix_word_token(m.group(0)), text)
    text = re.sub(r'(\d+)\s+0\.(\d+)', r'\1±0.\2', text)
    text = re.sub(r'(\d+)\s+MIN', r'\1 MIN', text, flags=re.I)
    text = re.sub(r'(\d+)\s+MAX', r'\1 MAX', text, flags=re.I)
    text = re.sub(r'WITHIN\s*([+-]?\d+)', r'WITHIN \1', text, flags=re.I)
    return text
def _correct_ocr_artifacts(t): return post_process_ocr(t)

def _rows_pdfplumber(b):
    rows = []; cat, last_item = "", ""
    try:
        with pdfplumber.open(io.BytesIO(b)) as pdf:
            for page in pdf.pages:
                for tbl in page.extract_tables():
                    for r in tbl:
                        raw = [(x or "").replace("\n"," ").strip() for x in r]
                        # Phase 2: collapse duplicated merged cells
                        c = []
                        for x in raw:
                            if not c or _norm(x) != _norm(c[-1]): c.append(x)
                        while c and not c[0]: c.pop(0)
                        if len(c) < 3:
                            if c and (_section_cat(c[0]) or _norm(c[0]) in _CATEGORIES): cat = c[0]
                            continue
                        sno, item, spec = c[0], c[1], c[2]
                        method = c[3] if len(c) > 3 else ""
                        tool = c[4] if len(c) > 4 else ""
                        sampling = c[5] if len(c) > 5 else ""
                        stage = c[6] if len(c) > 6 else ""
                        ni, ns = _norm(item), _norm(spec)
                        if ni in ("inspectionitem","item") or ns in ("criteria","spec","specification"): continue
                        chk = _norm(sno + item)
                        if chk and any(chk.startswith(s) or s in chk for s in _SKIP): continue
                        if ni in _CATEGORIES or _section_cat(item): cat = item; continue
                        if not spec: continue
                        if item: last_item = item
                        rows.append({"cat": cat, "item": _correct_ocr_artifacts(item or last_item),
                         "spec": _correct_ocr_artifacts(spec), "method": _correct_ocr_artifacts(method),
                         "sampling": _correct_ocr_artifacts(sampling), "stage": _correct_ocr_artifacts(stage),
                         "sno": _correct_ocr_artifacts(sno)})
    except Exception: pass
    return rows
def _rows_camelot(b):
    if not CAMELOT_AVAILABLE: return []
    rows = []
    try:
        for tbl in camelot.read_pdf(io.BytesIO(b), pages='all', flavor='lattice'):
            for _, row in tbl.df.iterrows():
                if len(row) < 3: continue
                sno, item, spec = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
                if not spec: continue
                rows.append({"cat":"","item":_correct_ocr_artifacts(item),"spec":_correct_ocr_artifacts(spec),
                 "method":_correct_ocr_artifacts(str(row[3]).strip() if len(row)>3 else ""),
                 "sampling":_correct_ocr_artifacts(str(row[4]).strip() if len(row)>4 else ""),
                 "stage":_correct_ocr_artifacts(str(row[5]).strip() if len(row)>5 else ""),
                 "sno":_correct_ocr_artifacts(sno)})
    except Exception: pass
    return rows
def _rows_tabula(b):
    if not TABULA_AVAILABLE: return []
    rows = []
    try:
        for df in tabula.read_pdf(io.BytesIO(b), pages='all', multiple_tables=True):
            for _, row in df.iterrows():
                cols = [str(x).strip() for x in row.values if pd.notna(x)]
                if len(cols) < 3: continue
                if not cols[2]: continue
                rows.append({"cat":"","item":_correct_ocr_artifacts(cols[1]),"spec":_correct_ocr_artifacts(cols[2]),
                 "method":_correct_ocr_artifacts(cols[3] if len(cols)>3 else ""),
                 "sampling":_correct_ocr_artifacts(cols[4] if len(cols)>4 else ""),
                 "stage":_correct_ocr_artifacts(cols[5] if len(cols)>5 else ""),
                 "sno":_correct_ocr_artifacts(cols[0])}) 
    except Exception: pass
    return rows
def _merge_rows(rows_list):
    seen = set(); merged = []
    for rows in rows_list:
        for r in rows:
            key = (_norm(r.get("item","")), _norm(r.get("spec","")))
            if key not in seen and _norm(r.get("spec","")): seen.add(key); merged.append(r)
    return merged
def _rows_pdfplumber_enhanced(b):
    res = _rows_pdfplumber(b)
    if len(res) < 5: res += _rows_camelot(b)
    if len(res) < 5: res += _rows_tabula(b)
    return _merge_rows([res])

try: _OCR_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".ocr_cache")
except NameError: _OCR_CACHE = os.path.join(os.getcwd(), ".ocr_cache")
@st.cache_resource(show_spinner=False)
def _rapidocr_engine():
    try:
        from rapidocr_onnxruntime import RapidOCR
        return RapidOCR()
    except Exception: return None
def preprocess_image_for_ocr(img):
    if not CV2_AVAILABLE: return img
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY) if len(img.shape) == 3 else img
    den = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    ang = detect_skew(den)
    if abs(ang) > 0.5: den = deskew_image(den, ang)
    enh = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8)).apply(den)
    return cv2.adaptiveThreshold(enh, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
def detect_skew(image):
    if not CV2_AVAILABLE: return 0.0
    coords = np.column_stack(np.where(image > 0))
    if len(coords) < 10: return 0.0
    a = cv2.minAreaRect(coords)[-1]
    return 90 + a if a < -45 else a
def deskew_image(image, angle):
    if not CV2_AVAILABLE: return image
    (h, w) = image.shape[:2]
    M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
    return cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
def get_optimal_ocr_dpi(image):
    if not CV2_AVAILABLE: return 220
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY) if len(image.shape) == 3 else image
    if cv2.Laplacian(gray, cv2.CV_64F).var() < 100: return 300
    return 250 if gray.std() < 30 else 200
def ocr_with_voting(image):
    allr = []
    try:
        e1 = _rapidocr_engine()
        if e1:
            res, _ = e1(image)
            if res: allr.append({"engine":"rapidocr","texts":[[bx,t,c] for bx,t,c in res]})
    except Exception: pass
    if TESSERACT_AVAILABLE and len(allr) < 2:
        try:
            d = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT); texts = []
            for i in range(len(d['text'])):
                if d['text'][i].strip():
                    texts.append([[d['left'][i], d['top'][i], d['left'][i]+d['width'][i], d['top'][i]+d['height'][i]],
                     d['text'][i].strip(), d['conf'][i]])
            allr.append({"engine":"tesseract","texts":texts})
        except Exception: pass
    if EASYOCR_AVAILABLE and len(allr) < 2:
        try:
            rd = easyocr.Reader(['en'], gpu=False); texts = []
            for box, t, c in rd.readtext(image, detail=1):
                xs = [p[0] for p in box]; ys = [p[1] for p in box]
                texts.append([[min(xs),min(ys),max(xs),max(ys)], t, c])
            allr.append({"engine":"easyocr","texts":texts})
        except Exception: pass
    if not allr: return []
    if len(allr) == 1: return allr[0]["texts"]
    pr = {"rapidocr":3,"tesseract":2,"easyocr":1}; best, bs = None, -1
    for r in allr:
        sc = pr.get(r.get("engine",""), 0); tx = r.get("texts", [])
        if tx: sc += sum(t[2] for t in tx)/len(tx)/100
        if sc > bs: bs, best = sc, r
    return best["texts"] if best else []
def _ocr_raw_pages(b, dpi=220):
    key = hashlib.md5(b).hexdigest() + f"_raw{dpi}"
    cf = os.path.join(_OCR_CACHE, key + ".json")
    if os.path.exists(cf):
        try:
            with open(cf, encoding="utf-8") as fh: return json.load(fh)
        except Exception: pass
    engine = _rapidocr_engine()
    if engine is None and not TESSERACT_AVAILABLE and not EASYOCR_AVAILABLE: return []
    doc = fitz.open(stream=b, filetype="pdf")
    use_multi = st.session_state.get("use_multi_ocr", False)
    def process_page(num, page):
        cur = dpi
        if dpi == 220 and CV2_AVAILABLE:
            prev = page.get_pixmap(matrix=fitz.Matrix(1,1), colorspace=fitz.csRGB)
            pimg = np.frombuffer(prev.samples, dtype=np.uint8).reshape(prev.h, prev.w, 3)
            od = get_optimal_ocr_dpi(pimg)
            if od > 220: cur = od
        pix = page.get_pixmap(matrix=fitz.Matrix(cur/72, cur/72), colorspace=fitz.csRGB)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, 3)
        if CV2_AVAILABLE: img = preprocess_image_for_ocr(img)
        if use_multi and (TESSERACT_AVAILABLE or EASYOCR_AVAILABLE): ocr_results = ocr_with_voting(img)
        else:
            try:
                res, _ = engine(img); ocr_results = [[bx,t,c] for bx,t,c in (res or [])]
            except Exception: ocr_results = []
        texts = []
        for item in ocr_results:
            if len(item) >= 3:
                box, txt, conf = item
                if isinstance(box, list) and box:
                    if isinstance(box[0], (int,float)):
                        if len(box) >= 4: x0,y0,x1,y1 = box[:4]
                        else: continue
                    else:
                        xs = [p[0] for p in box]; ys = [p[1] for p in box]
                        if xs and ys: x0,y0,x1,y1 = min(xs),min(ys),max(xs),max(ys)
                        else: continue
                else: continue
                t = str(txt).strip()
                if t: texts.append([(x0+x1)/2, (y0+y1)/2, y1-y0, t])
        return {"w": pix.w, "h": pix.h, "texts": texts}
    pages = [None]*len(doc)
    with ThreadPoolExecutor(max_workers=min(4, len(doc))) as ex:
        futs = {ex.submit(process_page, i, p): i for i, p in enumerate(doc)}
        for f in as_completed(futs): pages[futs[f]] = f.result()
    doc.close()
    try:
        os.makedirs(_OCR_CACHE, exist_ok=True)
        with open(cf, "w", encoding="utf-8") as fh: json.dump(pages, fh, ensure_ascii=False)
    except Exception: pass
    return pages

_BAND_HEADERS = {"sno":"sno","s.no":"sno","slno":"sno","inspectionitem":"item","item":"item",
 "criteria":"spec","specification":"spec","spec":"spec","inspectionmethod":"method","method":"method",
 "inspectiontool":"tool","tool":"tool","samplingplan":"sampling","sampling":"sampling",
 "inspectionstage":"stage","stage":"stage"}
_MIS_BOUNDS = [(0.00,0.14,"sno"),(0.14,0.33,"item"),(0.33,0.455,"spec"),(0.455,0.555,"method"),
 (0.555,0.685,"tool"),(0.685,0.775,"sampling"),(0.775,1.00,"stage")]
_MIS_META = ("partname","partnumber","vendorcode","vendorsname","vendor'sname","model","batch",
 "code","issueno","issuedate","standard","typeof")
_DIM_SPEC_RE = re.compile(r"\d[\d.\s]*\s*(?:±|\+\s*/?\s*-)\s*\.?\d|\d[\d.]*\s*~\s*\.?\d")
_ENUM_RE = re.compile(r"^\s*(?:\(?\s*(?:\d{1,3}|[A-Za-z]|[ivxIVX]{1,4})\s*[\.\)\]]?\s+|\d{1,3}(?=[A-Z]))")
_ENUM_ONLY = re.compile(r"^[\(\[]?\s*(?:\d{1,3}(?:\.\d{1,2})?|[a-eA-E]|[ivxlIVXL]{1,4}|川)\s*[\.\)\]]?$")
_FOOTER_STRONG = ("approvedby","inspectedby","verifiedby","formno","documentclassification","revisionnumber","qaqa")
_FOOTER_WEAK = ("approved","checked","prepared")
def _is_mis_form(texts): return any("inspectionstandard" in _norm(t) for _,_,_,t in texts)
def _mis_body_start(texts, page_h):
    ys = [y for x,y,h,t in texts if _norm(t).startswith(_MIS_META) and y < page_h*0.22]
    return max(ys) if ys else page_h*0.21
def _footer_cut(texts, page_h):
    hits = []
    for x,y,h,t in texts:
        n = _norm(t)
        if n.startswith(_FOOTER_STRONG): hits.append((y, True))
        elif n.startswith(_FOOTER_WEAK): hits.append((y, False))
    cut = page_h
    for y, strong in hits:
        if strong or any(abs(y-y2) < 130 for y2,_ in hits): cut = min(cut, y)
    return cut
def _strip_enum(s):
    prev = None; s = s.strip()
    while s != prev: prev = s; s = _ENUM_RE.sub("", s, count=1).strip()
    return s
def _despace(text):
    try: import wordninja
    except Exception: return text
    out = []
    for tok in str(text).split():
        if tok.isalpha() and len(tok) >= 15:
            parts = wordninja.split(tok)
            if len(parts) >= 2 and all(len(p) >= 2 for p in parts): out.append(" ".join(parts)); continue
        out.append(tok)
    return " ".join(out)
def _rows_bands(b, dpi=220):
    rows = []; cat = ""
    for pg in _ocr_raw_pages(b, dpi):
        texts = [tuple(t) for t in pg["texts"]]
        if not texts: continue
        header_hits = [(x,y,h,t) for x,y,h,t in texts if _norm(t) in _BAND_HEADERS]
        roles_x, hdr_y, bounds = {}, None, None
        if header_hits:
            hdr_y = sorted(h[1] for h in header_hits)[len(header_hits)//2]
            hdr = [(x, _BAND_HEADERS[_norm(t)]) for x,y,h,t in header_hits if abs(y-hdr_y) < 40]
            for x, role in sorted(hdr): roles_x.setdefault(role, x)
            if "spec" in roles_x and "item" in roles_x:
                ordered = sorted(roles_x.items(), key=lambda kv: kv[1]); bounds = []
                for k in range(len(ordered)):
                    x0 = 0 if k == 0 else (ordered[k-1][1]+ordered[k][1])/2
                    x1 = pg["w"] if k == len(ordered)-1 else (ordered[k][1]+ordered[k+1][1])/2
                    bounds.append((x0,x1,ordered[k][0]))
        if bounds is None and _is_mis_form(texts):
            w = pg["w"]; bounds = [(a*w,b*w,r) for a,b,r in _MIS_BOUNDS]
            roles_x = {r:(a+b)/2*w for a,b,r in _MIS_BOUNDS}; hdr_y = _mis_body_start(texts, pg["h"])
        if bounds is None or "spec" not in roles_x or "item" not in roles_x: continue
        cut_y = _footer_cut(texts, pg["h"])
        def role_of(x):
            for x0,x1,r in bounds:
                if x0 <= x < x1: return r
            return None
        band = {"item":[],"spec":[],"method":[],"sampling":[],"stage":[]}
        for x,y,h,t in texts:
            if not (hdr_y+18 < y < cut_y): continue
            r = role_of(x)
            if r == "sno" and not _ENUM_ONLY.match(str(t).strip()): r = "item"
            if r in band: band[r].append((x,y,h,t))
        if not band["item"]: continue
        allf = band["item"]+band["spec"]+band["method"] or band["item"]
        medh = sorted(f[2] for f in allf)[len(allf)//2] if allf else 20.0
        medh = medh if medh > 4 else 20.0
        gaps = []
        for col in (band["item"],band["spec"],band["method"]):
            ys = sorted(f[1] for f in col)
            gaps += [b-a for a,b in zip(ys,ys[1:]) if medh*0.5 < b-a < medh*1.5]
        pitch = sorted(gaps)[len(gaps)//2] if gaps else medh*1.3
        row_gap = min(pitch*1.9+4, medh*1.5+4)
        def cluster(frags, breaks=()):
            frags = sorted(frags, key=lambda v: v[1]); blocks, cur, last = [], [], None
            for f in frags:
                crossed = last is not None and any(last < bb <= f[1] for bb in breaks)
                if last is not None and (f[1]-last > row_gap or crossed): blocks.append(cur); cur = []
                cur.append(f); last = f[1]
            if cur: blocks.append(cur)
            out = []
            for bl in blocks:
                y0 = min(f[1] for f in bl); y1 = max(f[1] for f in bl)
                txt = " ".join(f[3] for f in sorted(bl, key=lambda v:(round(v[1]/(pitch/2+1)), v[0])))
                out.append((y0,y1,re.sub(r"\s+"," ",txt).strip()))
            return out
        def gather(frags, y0, y1, pad):
            sel = [f for f in frags if y0-pad <= f[1] <= y1+pad]
            return re.sub(r"\s+"," "," ".join(f[3] for f in sorted(sel, key=lambda v:(round(v[1]/(pitch/2+1)), v[0])))).strip()
        item_blocks = cluster(band["item"]); item_x = []
        for iy0,iy1,_ in item_blocks:
            xs = [f[0] for f in band["item"] if iy0 <= f[1] <= iy1 and not _ENUM_ONLY.match(str(f[3]).strip())]
            item_x.append(min(xs) if xs else 1e9)
        med_x = sorted(item_x)[len(item_x)//2] if item_x else 0.0
        item_cy = sorted((iy0+iy1)/2 for iy0,iy1,_ in item_blocks)
        item_breaks = [(item_cy[k]+item_cy[k+1])/2 for k in range(len(item_cy)-1)]
        spec_blocks = cluster(band["spec"], breaks=item_breaks)
        def overlaps_spec(iy0,iy1):
            return any(not (sy1 < iy0-pitch or sy0 > iy1+pitch) for sy0,sy1,_ in spec_blocks)
        headers, prefix_of, active_group = [], {}, None
        for k,(iy0,iy1,itext) in enumerate(item_blocks):
            itc = _strip_enum(itext); has_spec = overlaps_spec(iy0,iy1)
            canon = _section_cat(itc) or ((_norm(itc) in _CATEGORIES) and itc.upper())
            if (not has_spec) and canon:
                headers.append(((iy0+iy1)/2, canon if isinstance(canon,str) else itc)); active_group = None
            elif item_x[k] >= med_x-8: active_group = itc if not has_spec else None
            elif has_spec and active_group: prefix_of[k] = active_group
        headers.sort()
        def cat_at(y):
            c = ""
            for hy,ht in headers:
                if hy <= y: c = ht
                else: break
            return c
        def nearest_item_k(sc):
            best, bestd = None, 1e9
            for k,(iy0,iy1,_) in enumerate(item_blocks):
                d = 0 if (iy0-pitch <= sc <= iy1+pitch) else min(abs(sc-iy0), abs(sc-iy1))
                if d < bestd: bestd, best = d, k
            return best if bestd < pitch*3.5 else None
        pstart = len(rows)
        for sy0,sy1,stext in spec_blocks:
            sc = (sy0+sy1)/2; ki = nearest_item_k(sc)
            base = _strip_enum(item_blocks[ki][2]) if ki is not None else ""
            pref = prefix_of.get(ki,"")
            itext = _despace(f"{_strip_enum(pref)} {base}".strip() if pref else base)
            if not _norm(itext+stext): continue
            rows.append({"cat":cat_at(sc),"item":_correct_ocr_artifacts(itext),"spec":_correct_ocr_artifacts(stext),
             "method":_correct_ocr_artifacts(gather(band["method"],sy0,sy1,pitch*0.6)),
             "sampling":_correct_ocr_artifacts(gather(band["sampling"],sy0,sy1,pitch*0.6)),
             "stage":_correct_ocr_artifacts(gather(band.get("stage",[]),sy0,sy1,pitch*0.6)),"sno":""})
        emitted = " ".join(_norm(r["spec"]) for r in rows[pstart:])
        for x,y,hh,t in sorted(band["spec"], key=lambda v: v[1]):
            if not _DIM_SPEC_RE.search(t): continue
            nt = _norm(t)
            if nt and nt in emitted: continue
            ki = nearest_item_k(y)
            itext = _despace(_strip_enum(item_blocks[ki][2])) if ki is not None else ""
            rows.append({"cat":cat_at(y),"item":_correct_ocr_artifacts(itext),"spec":_correct_ocr_artifacts(t.strip()),
             "method":_correct_ocr_artifacts(gather(band["method"],y,y,pitch*0.9)),
             "sampling":_correct_ocr_artifacts(gather(band["sampling"],y,y,pitch*0.9)),
             "stage":_correct_ocr_artifacts(gather(band.get("stage",[]),y,y,pitch*0.9)),"sno":""})
            emitted += " " + nt
    return [r for r in rows if _norm(r["item"]+r["spec"]) and not any(_norm(r["item"]).startswith(s) for s in _SKIP)]
@st.cache_resource(show_spinner="Loading Docling…")
def _docling_converter(do_ocr, fast=False):
    try:
        from docling.document_converter import DocumentConverter, PdfFormatOption
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.datamodel.base_models import InputFormat
        opts = PdfPipelineOptions(); opts.do_ocr = do_ocr; opts.do_table_structure = True
        opts.table_structure_options.do_cell_matching = True
        try:
            from docling.datamodel.pipeline_options import TableFormerMode
            opts.table_structure_options.mode = TableFormerMode.FAST if fast else TableFormerMode.ACCURATE
        except Exception: pass
        return DocumentConverter(format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=opts)})
    except Exception: return None
def _rows_docling(b, do_ocr=True, fast=False):
    try:
        from docling.datamodel.base_models import DocumentStream
        conv = _docling_converter(do_ocr, fast)
        if conv is None: return []
        doc = conv.convert(DocumentStream(name="doc.pdf", stream=io.BytesIO(b))).document
    except Exception: return []
    rows = []; cat, last_item = "", ""
    for tb in doc.tables:
        try: df = tb.export_to_dataframe(doc)
        except Exception: continue
        for i in range(len(df)):
            raw = [str(x).replace("\n"," ").strip() for x in df.iloc[i].tolist()]
            c = []
            for x in raw:
                if not c or _norm(x) != _norm(c[-1]): c.append(x)
            if len(c) < 3: continue
            sno, item, spec = c[0], c[1], c[2]
            chk = _norm(sno+item)
            if chk and any(chk.startswith(s) or s in chk for s in _SKIP): continue
            if _norm(item) in _CATEGORIES: cat = item; continue
            if _norm(spec) == _norm(item) or not spec: continue
            if item: last_item = item
            rows.append({"cat":cat,"item":_correct_ocr_artifacts(item or last_item),"spec":_correct_ocr_artifacts(spec),
             "method":_correct_ocr_artifacts(c[3] if len(c)>3 else ""),
             "sampling":_correct_ocr_artifacts(c[5] if len(c)>5 else (c[4] if len(c)>4 else "")),
             "stage":_correct_ocr_artifacts(c[6] if len(c)>6 else ""),"sno":_correct_ocr_artifacts(sno)})
    return rows
@st.cache_data(show_spinner="Reading PDF…")
def extract_pdf_rows(b, fast=False):
    if _digital_words_per_page(b) >= 25:
        return _rows_pdfplumber_enhanced(b), "digital · pdfplumber+"
    mode = "fast" if fast else "accurate"
    cf = os.path.join(_OCR_CACHE, hashlib.md5(b).hexdigest()+"_bands10_"+mode+".json")
    if os.path.exists(cf):
        try:
            with open(cf, encoding="utf-8") as fh: d = json.load(fh)
            return d["rows"], d["src"]+" · cached"
        except Exception: pass
    rows = _rows_bands(b); src = "scanned · RapidOCR column-bands (PP-OCRv6) + Phase4A"
    if len(rows) < 5:
        try:
            rows = _rows_docling(b, do_ocr=True, fast=fast)
            src = f"scanned · Docling + RapidOCR (PP-OCRv6, {mode}) + Phase4A"
        except Exception: rows = rows or []
    if not rows: return [], "scanned · OCR unavailable (read the image below)"
    try:
        os.makedirs(_OCR_CACHE, exist_ok=True)
        with open(cf, "w", encoding="utf-8") as fh: json.dump({"rows":rows,"src":src}, fh, ensure_ascii=False)
    except Exception: pass
    return rows, src

# ════ 5. METADATA EXTRACTION (Phase 1) ════
_MSIL_VENDOR_RE = re.compile(r"([A-Z])\s*(\d)\s*(\d)\s*(\d)(?:\s*(\d))?")
_MSIL_VENDOR_COMPACT_RE = re.compile(r"([A-Z])(\d{3,4})\b")
_MSIL_PARTNO_RE = re.compile(r"\b(\d{5}(?:M\d{2}[A-Z0-9]{2,}|-?\d{2}[A-Z0-9]{2,}))\b", re.IGNORECASE)
_META_EMPTY = {"vendor_code": None, "part_number": None, "model_no": None}
_VENDOR_OCR_DIGIT = str.maketrans({"O":"0","Q":"0","I":"1","|":"1","S":"5","B":"8","Z":"2"})
_VENDOR_FIRST_DIGIT = {"5":"S","0":"O","1":"I","8":"B","2":"Z"}
_HEADER_DATE_RE = re.compile(r"(\d{1,2})\s*[./\-]\s*(\d{1,2})\s*[./\-]\s*(\d{2,4})")
def _clean_cell(c): return re.sub(r"\s+"," ", str(c or "").replace("\n"," ")).strip()
def _vendor_from_row(cells):
    run, best = [], None
    def flush(r):
        if len(r) >= 4 and all(x.isdigit() for x in r[1:4]):
            first = r[0] if r[0].isalpha() else _VENDOR_FIRST_DIGIT.get(r[0])
            if first and first.isalpha():
                code = first + "".join(r[1:4])
                if len(r) >= 5 and r[4].isdigit(): code += r[4]
                return code
        return None
    for c in cells:
        raw = re.sub(r"[^A-Za-z0-9]","", str(c))
        if len(raw) == 1: run.append(raw.upper())
        else:
            best = best or flush(run); run = []
            if 4 <= len(raw) <= 5:
                m = _MSIL_VENDOR_COMPACT_RE.fullmatch(raw)
                if m: best = best or m.group(1)
    return best or flush(run)
def _part_from_row(cells):
    for c in cells:
        m = _MSIL_PARTNO_RE.search(re.sub(r"\s+","", str(c)))
        if m: return m.group(1).upper()
    return None
def _model_from_row(cells):
    for i, c in enumerate(cells):
        n = _norm(c)
        if n == "model":
            for cc in cells[i+1:]:
                nn = _norm(cc)
                if not nn or nn in ("model","partname","partnumber","batch","batchcode") or nn.startswith(("part","model")): continue
                return _clean_cell(cc)
        elif n.startswith("model") and len(n) > 6:
            v = re.sub(r"^model[\s:.\-]*","", str(c), flags=re.I)
            if _norm(v): return _clean_cell(v)
    return None
def _date_from_row(cells):
    for c in cells:
        c2 = str(c).replace("：",".").replace(":",".")
        m = _HEADER_DATE_RE.search(c2)
        if m and 1 <= int(m.group(1)) <= 31 and 1 <= int(m.group(2)) <= 12:
            return f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.{m.group(3)}"
    return None
def _meta_from_rows(tables):
    meta = dict(_META_EMPTY); meta["issue_date"] = None
    for tbl in tables:
        rows = [[_clean_cell(c) for c in r] for r in tbl]
        prev_issue = False
        for row in rows[:8]:
            norms = [_norm(c) for c in row]
            has_issue = any(n == "issuedate" or n.startswith("issuedate") for n in norms)
            if not meta["vendor_code"]: meta["vendor_code"] = _vendor_from_row(row)
            if not meta["part_number"]: meta["part_number"] = _part_from_row(row)
            if not meta["model_no"]: meta["model_no"] = _model_from_row(row)
            if not meta["issue_date"] and (has_issue or prev_issue): meta["issue_date"] = _date_from_row(row)
            prev_issue = has_issue
    return meta
def _ocr_box_rows(texts, tol=14):
    rows = []
    for x,y,h,t in sorted(texts, key=lambda z:(z[1],z[0])):
        for row in rows:
            if abs(row[0]-y) < tol: row[1].append((x,t)); break
        else: rows.append((y,[(x,t)]))
    return [[t for _,t in sorted(cs, key=lambda z:z[0])] for _,cs in sorted(rows, key=lambda z:z[0])]
def _parse_msil_vendor_code(text):
    if not text: return None
    up = str(text).upper(); m = _MSIL_VENDOR_RE.search(up)
    if m:
        code = m.group(1)+m.group(2)+m.group(3)+m.group(4)
        if m.group(5): code += m.group(5)
        return code
    compact = re.sub(r"\s+","", up).translate(_VENDOR_OCR_DIGIT)
    m = _MSIL_VENDOR_COMPACT_RE.search(compact)
    return m.group(1)+m.group(2) if m else None
def _vendor_code_from_boxes(boxes):
    if not boxes: return None
    boxes = sorted(boxes, key=lambda b:(b[1],b[0]))
    code = _parse_msil_vendor_code(" ".join(str(b[3]) for b in boxes))
    if code: return code
    lines = []
    for b in boxes:
        for line in lines:
            if abs(line[0][1]-b[1]) < max(12, b[2]*0.6): line.append(b); break
        else: lines.append([b])
    for line in lines:
        line.sort(key=lambda b:b[0]); chars = []
        for _x,_y,_h,t in line:
            raw = re.sub(r"[^A-Z0-9]","", str(t).upper())
            if not raw: continue
            if len(raw) == 1:
                c = raw.translate(_VENDOR_OCR_DIGIT) if raw.isdigit() or raw in "O0Q" else raw
                if c.isalpha() or c.isdigit(): chars.append(c)
            elif re.match(r"^[A-Z]\d{3,4}$", raw): return raw
            else:
                p = _parse_msil_vendor_code(raw)
                if p: return p
        if len(chars) >= 4 and chars[0].isalpha() and all(c.isdigit() for c in chars[1:4]):
            code = chars[0]+"".join(chars[1:4])
            if len(chars) >= 5 and chars[4].isdigit(): code += chars[4]
            return code
    return None
def _is_vendor_header_label(t):
    n = _norm(t)
    if n in ("vendor","vendorcode"): return True
    if "vendor" not in n: return False
    return not any(s in n for s in ("name","mark","batch","partno","standard"))
def _find_vendor_header_box(texts, page_h):
    hits = [(x,y,h,t) for x,y,h,t in texts if _is_vendor_header_label(t) and y < page_h*0.22]
    if not hits: return None
    exact = [b for b in hits if _norm(b[3]) == "vendor"]
    return min((exact or hits), key=lambda b:(b[1],b[0]))
def _find_code_label_box(texts, vx, vy):
    for x,y,h,t in texts:
        if _norm(t) == "code" and abs(x-vx) < 50 and 5 < y-vy < 55: return (x,y,h,t)
    return None
def _vendor_value_column_bounds(vx, vy, code_box, texts, page_w, page_h):
    rh = [x for x,y,h,t in texts if x > vx+15 and y <= vy+70 and _norm(t).startswith(("typeof","standard","issue","date","page","issueno"))]
    x_right = (vx+min(rh))/2 if rh else vx+page_w*0.06
    half = x_right-vx; x_left = vx-half*0.85
    y_top = code_box[1]+code_box[2]*0.45 if code_box else vy+30
    return x_left, x_right, y_top, min(vy+100, page_h*0.16)
def _extract_vendor_code_spatial(texts, page_w, page_h):
    header = _find_vendor_header_box(texts, page_h)
    if not header: return None
    vx, vy = header[0], header[1]
    cb = _find_code_label_box(texts, vx, vy)
    x0,x1,y0,y1 = _vendor_value_column_bounds(vx, vy, cb, texts, page_w, page_h)
    skip = {"vendor","vendorcode","code"}
    vb = [b for b in texts if x0 <= b[0] <= x1 and y0 <= b[1] <= y1 and _norm(b[3]) not in skip]
    return _vendor_code_from_boxes(vb)
def _find_label_box(texts, keys, page_h, y_frac=0.25):
    hits = []
    for x,y,h,t in texts:
        if y > page_h*y_frac: continue
        n = _norm(t)
        for k in keys:
            if n == k or (k in n and len(n) <= len(k)+6): hits.append((len(k),y,x,(x,y,h,t)))
    if not hits: return None
    hits.sort(key=lambda z:(-z[0],z[1],z[2]))
    return hits[0][3]
def _is_header_edge_label(t):
    return _norm(t) in {"partname","partnumber","vendor","vendorcode","vendorsname","batch","code",
     "typeof","standard","issue","date","page","mis","model","issueno","applicable","inspection","sampling","remarks"}
def _neighbor_header_x(lx, ly, texts, page_w, side="right", y_tol=25):
    xs = [x for x,y,h,t in texts if y <= ly+80 and _is_header_edge_label(t)
     and (x > lx+20 if side == "right" else x < lx-20) and abs(y-ly) <= y_tol]
    if not xs: return lx+page_w*(0.08 if side == "right" else -0.08)
    return (lx+(min(xs) if side == "right" else max(xs)))/2
def _field_value_bounds(label, texts, page_w, page_h, mode="below"):
    lx, ly, lh, _lt = label
    if mode == "right":
        return lx+lh*0.4, _neighbor_header_x(lx,ly,texts,page_w,"right"), ly-lh*0.8, ly+lh*1.1
    xl = _neighbor_header_x(lx,ly,texts,page_w,"left",30)
    xr = max(_neighbor_header_x(lx,ly,texts,page_w,"right",30), min(page_w*0.92, lx+page_w*0.20))
    return xl, xr, ly+lh*0.1, min(ly+55, page_h*0.19)
def _join_meta_boxes(boxes, skip=None):
    parts = []
    for _x,_y,_h,t in sorted(boxes, key=lambda b:(b[1],b[0])):
        raw = str(t).strip()
        if not raw: continue
        n = _norm(raw)
        if skip and (n in skip or any(n.startswith(s) for s in skip)): continue
        parts.append(raw)
    return re.sub(r"\s+"," "," ".join(parts)).strip()
def _parse_model_no(text):
    if not text: return None
    up = re.sub(r"\s+"," ", str(text)).strip().upper()
    if not up or _norm(up) in ("model","no","modelno"): return None
    up = re.sub(r"[^A-Z0-9 -/.]"," ", up); up = re.sub(r"\s+"," ", up).strip()
    if len(up) >= 2 and not up.startswith("AS PER"): return up.split()[0] if up.split() else None
    return None
def _parse_part_number(text):
    if not text: return None
    raw = re.sub(r"\s+"," ", str(text)).strip()
    m = _MSIL_PARTNO_RE.search(re.sub(r"\s+","", raw.upper()))
    if m: return m.group(1).upper()
    cleaned = raw
    for junk in ("PART NUMBER:","PART NUMBER","PART NO:","PART NO"):
        cleaned = re.sub(re.escape(junk)," ",cleaned,flags=re.I)
    cleaned = re.sub(r"\s+"," ",cleaned).strip(); n = _norm(cleaned)
    if not cleaned or n in ("part","number","partnumber","name","partname"): return None
    if cleaned.upper().startswith("AS PER") or len(cleaned) >= 3: return cleaned
    return None
def _boxes_in_region(texts, x0, x1, y0, y1, skip=None):
    return [(x,y,h,t) for x,y,h,t in texts if x0 <= x <= x1 and y0 <= y <= y1
     and not (skip and _norm(t) in skip)]
def _extract_model_spatial(texts, page_w, page_h):
    label = _find_label_box(texts, ("model",), page_h)
    if not label: return None
    x0,x1,y0,y1 = _field_value_bounds(label, texts, page_w, page_h, "right")
    return _parse_model_no(_join_meta_boxes(_boxes_in_region(texts,x0,x1,y0,y1,{"model","no","modelno"}),{"model","no","modelno"}))
def _extract_part_number_spatial(texts, page_w, page_h):
    label = _find_label_box(texts, ("partnumber",), page_h)
    if not label: return None
    x0,x1,y0,y1 = _field_value_bounds(label, texts, page_w, page_h, "below")
    skip = {"part","number","partnumber","name","partname"}
    return _parse_part_number(_join_meta_boxes(_boxes_in_region(texts,x0,x1,y0,y1,skip),skip))
def extract_metadata_from_pdf(b):
    meta = dict(_META_EMPTY); meta["issue_date"] = None
    # 1) digital tables
    try:
        with pdfplumber.open(io.BytesIO(b)) as pdf:
            tables = [t for page in pdf.pages[:4] for t in (page.extract_tables() or [])]
        m = _meta_from_rows(tables)
        for k, v in m.items():
            if not meta.get(k) and v: meta[k] = v
    except Exception: pass
    # 2) OCR rows
    if not all(meta.get(k) for k in _META_EMPTY):
        try:
            for pg in _ocr_raw_pages(b, 220)[:4]:
                m = _meta_from_rows([_ocr_box_rows(pg["texts"])])
                for k, v in m.items():
                    if not meta.get(k) and v: meta[k] = v
                if all(meta.get(k) for k in _META_EMPTY): break
        except Exception: pass
    # 3) spatial fallback
    if not all(meta.get(k) for k in _META_EMPTY):
        try:
            pages = _ocr_raw_pages(b, 220)
            if pages and pages[0]["texts"]:
                pg = pages[0]; texts, w, h = pg["texts"], pg["w"], pg["h"]
                if not meta["vendor_code"]: meta["vendor_code"] = _extract_vendor_code_spatial(texts,w,h)
                if not meta["model_no"]: meta["model_no"] = _extract_model_spatial(texts,w,h)
                if not meta["part_number"]: meta["part_number"] = _extract_part_number_spatial(texts,w,h)
        except Exception: pass
    return meta

# ════ 6. TOLERANCE ENGINE ════
_NUM = r"[-+]?\d+(?:\.\d+)?"
def _decimals(s): s = str(s); return len(s.split(".")[1]) if "." in s else 0
def _min_decimals(x, cap=4):
    if x is None: return 0
    s = f"{round(float(x),cap):.{cap}f}".rstrip("0").rstrip(".")
    return len(s.split(".")[1]) if "." in s else 0
def _fmt_num(x, d):
    v = round(float(x), d); return int(v) if v == int(v) else v
def _dev_result(nom, devs):
    t = float(nom); vals = [float(d.replace(" ","")) for d in devs]
    if len(vals) == 1: vals.append(0.0)
    lo, hi = t+min(vals), t+max(vals)
    if lo == hi: return None
    return {"target":t,"lower":lo,"upper":hi,"decimals":max([_decimals(nom)]+[_decimals(d) for d in devs]),"has_target":True}
_UOM_PATTERN = re.compile(r'(?P<uom>mm|MPa|Shore\s*A|%|g/cm³|g/cc|N/mm²|N|Kpa|°C|gm/cc|Nm|N/m|HRC|RPM|V|A|dB|Kg|kg|LPM|cm|m|g|ml|l|°|K|J|W|Hz|Pa|bar|psi|torr|atm|mmHg|H|D|B|S|T)\s*$', re.IGNORECASE)
def extract_uom(text):
    if not text: return None
    m = _UOM_PATTERN.search(str(text).strip())
    return m.group('uom') if m else None
def parse_spec(text):
    if text is None: return None
    s = str(text).strip().replace("−","-")
    if not s: return None
    uom = extract_uom(s)
    if uom: s = re.sub(rf'\s*{re.escape(uom)}\s*$','',s).strip()
    m = re.search(rf"({_NUM})\s*±\s*({_NUM})", s)
    if m:
        t, tol = float(m.group(1)), float(m.group(2))
        return {"target":t,"lower":t-tol,"upper":t+tol,"decimals":max(_decimals(m.group(1)),_decimals(m.group(2))),"has_target":True,"uom":uom}
    m = re.search(rf"({_NUM})\s*\+\s*({_NUM})\s*/?\s*-\s*({_NUM})", s)
    if m:
        t,a,b = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return {"target":t,"lower":t-b,"upper":t+a,"decimals":max(_decimals(m.group(1)),_decimals(m.group(2)),_decimals(m.group(3))),"has_target":True,"uom":uom}
    m = re.search(rf"({_NUM})\s*([+-]?\s*\d[\d.]*)\s*/\s*([+-]\s*\d[\d.]*)", s)
    if m:
        r = _dev_result(m.group(1), [m.group(2), m.group(3)])
        if r: r["uom"] = uom; return r
    mm = re.search(rf"({_NUM})\s*MIN", s, re.I)
    if mm:
        v = float(mm.group(1))
        return {"target":None,"lower":v,"upper":None,"decimals":_min_decimals(v),"has_target":False,"uom":uom}
    mx = re.search(rf"({_NUM})\s*MAX", s, re.I)
    if mx:
        v = float(mx.group(1))
        return {"target":None,"lower":None,"upper":v,"decimals":_min_decimals(v),"has_target":False,"uom":uom}
    m = re.match(rf"\s*[øØφΦ⌀]?\s*(?:[oO0]\s+)?({_NUM})\s*(?:~|to)\s*({_NUM})", s)
    if m:
        lo, hi = float(m.group(1)), float(m.group(2))
        if lo != hi:
            lo, hi = min(lo,hi), max(lo,hi)
            return {"target":(lo+hi)/2,"lower":lo,"upper":hi,"decimals":max(_decimals(m.group(1)),_decimals(m.group(2))),"has_target":False,"uom":uom}
    m = re.match(rf"\s*({_NUM})\s+([+-]\s*\d[\d.]*)\s*[a-zA-Zµ°%]{{0,4}}\s*$", s)
    if m:
        r = _dev_result(m.group(1), [m.group(2)])
        if r: r["uom"] = uom; return r
    return None
def parse_limits(param):
    spec = parse_spec(param)
    if not spec: return None
    d = max(_min_decimals(spec.get("lower")), _min_decimals(spec.get("upper")))
    return {"lower":spec.get("lower"),"upper":spec.get("upper"),
     "target":spec.get("target") if spec.get("has_target") else None,"decimals":d,"uom":spec.get("uom")}

# ════ 7. AUTOMATION PIPELINE ════
def _row_has_inspection_content(row):
    return str(row["Inspection Item"]).strip().lower() not in ("","nan","none") or str(row["Parameter"]).strip().lower() not in ("","nan","none")
def auto_fill_long_text(df):
    lt = df["Inspection Item"].fillna("").astype(str)+" "+df["Parameter"].fillna("").astype(str)+" "+df["Info Field 3"].fillna("").astype(str)
    df["Long Text"] = lt.str.replace(r'\s+',' ',regex=True).str.strip()
    return df
def auto_fix_issue_date(df):
    for i in range(len(df)):
        val = df.iloc[i]["Issue date"]
        if pd.isna(val): continue
        if isinstance(val, (datetime.datetime, datetime.date)):
            try: df.at[df.index[i],"Issue date"] = val.strftime("%d-%m-%Y"); continue
            except Exception: pass
        s = str(val).strip()
        if not s or s.lower() in ("nan","none","nat"): continue
        if ' ' in s: s = s.split(' ')[0]
        if 'T' in s: s = s.split('T')[0]
        try:
            dt = pd.to_datetime(s, errors='coerce')
            if pd.notna(dt): df.at[df.index[i],"Issue date"] = dt.strftime("%d-%m-%Y"); continue
        except Exception: pass
        cleaned = re.sub(r'[/\\._ ]','-',s)
        if re.match(r'^\d{2}-\d{2}-\d{4}$', cleaned): df.at[df.index[i],"Issue date"] = cleaned; continue
        if re.match(r'^\d{4}-\d{2}-\d{2}$', cleaned):
            p = cleaned.split('-'); df.at[df.index[i],"Issue date"] = f"{p[2]}-{p[1]}-{p[0]}"
    return df
def auto_number_operations(df):
    c = 1
    for i in range(len(df)):
        if _row_has_inspection_content(df.iloc[i]): df.at[df.index[i],"Operation number"] = c; c += 1
    return df
def auto_calculate_limits(df):
    for i in range(len(df)):
        param = df.iloc[i]["Parameter"]
        if pd.isna(param): continue
        lim = parse_limits(param)
        if lim:
            d = lim["decimals"]
            df.at[df.index[i],"Lower Limit"] = _fmt_num(lim["lower"],d) if lim["lower"] is not None else None
            df.at[df.index[i],"Target Value"] = None
            df.at[df.index[i],"Upper Limit"] = _fmt_num(lim["upper"],d) if lim["upper"] is not None else None
            df.at[df.index[i],"Decimal Places"] = d if d > 0 else None
            df.at[df.index[i],"UOM"] = lim.get("uom")
    return df
def auto_fix_mic_pattern(df):
    fixed = df.copy(); last = None
    for i in range(len(fixed)):
        val = fixed.iloc[i]["MIC Name"]
        if pd.isna(val) or str(val).strip() == "": fixed.at[fixed.index[i],"MIC Name"] = last
        else: last = str(val).strip()
    return fixed
def auto_clean_decimal_places(df):
    def clean(v):
        if pd.isna(v): return None
        s = str(v).strip()
        if s == "": return None
        try:
            f = float(s)
            if f == 0: return None
            return int(f) if f == int(f) else f
        except ValueError: return v
    df["Decimal Places"] = df["Decimal Places"].apply(clean)
    return df
def _auto_summary(df):
    return {"long_text":sum(1 for x in df["Long Text"] if str(x).strip()),
     "issue_date":sum(1 for x in df["Issue date"] if pd.notna(x) and str(x).strip()),
     "operation_no":sum(_row_has_inspection_content(df.iloc[i]) for i in range(len(df))),
     "limits":sum(1 for i in range(len(df)) if parse_limits(df.iloc[i]["Parameter"]) is not None),
     "mic_pattern":sum(1 for x in df["MIC Name"] if pd.notna(x) and str(x).strip()),
     "decimal_places":sum(1 for x in df["Decimal Places"] if pd.notna(x))}
def run_automation_pipeline(df):
    for fn in (auto_fill_long_text,auto_fix_issue_date,auto_number_operations,
     auto_calculate_limits,auto_fix_mic_pattern,auto_clean_decimal_places): df = fn(df)
    return df, _auto_summary(df)
def reapply_automation(df):
    orig = df.copy()
    df, summ = run_automation_pipeline(df)
    fixed = 0
    for col in ["Long Text","Issue date","Operation number","Lower Limit","Target Value","Upper Limit","Decimal Places","MIC Name"]:
        if col in df.columns and col in orig.columns:
            fixed += (df[col].astype(str) != orig[col].astype(str)).sum()
    return df, summ, fixed

# ════ 8. VALIDATION ENGINE (+ Phase 3) ════
def validate_required_cells(df):
    out = []
    for i in range(len(df)):
        for col in REQUIRED_COLS:
            if col in df.columns:
                v = df.iloc[i][col]
                if pd.isna(v) or str(v).strip() == "":
                    out.append({"row":i+1,"row_index":i,"column":col,"issue":"Cell is empty - must fill","severity":"critical"})
    return out
def validate_metadata_match(df, pv, pp, pm):
    out = []
    for i in range(len(df)):
        r = df.iloc[i]
        ev, ep, em = str(r.get("VENDOR CODE","")).strip(), str(r.get("Part number","")).strip(), str(r.get("Model No.","")).strip()
        if pv and ev and ev != pv: out.append({"row":i+1,"row_index":i,"column":"VENDOR CODE","issue":f"PDF ({pv}) ≠ Excel ({ev})","severity":"warning"})
        if pp and ep and ep != pp: out.append({"row":i+1,"row_index":i,"column":"Part number","issue":f"PDF ({pp}) ≠ Excel ({ep})","severity":"warning"})
        if pm and em and em != pm: out.append({"row":i+1,"row_index":i,"column":"Model No.","issue":f"PDF ({pm}) ≠ Excel ({em})","severity":"warning"})
    return out
def validate_decimal_places(df):
    out = []
    for i in range(len(df)):
        v = df.iloc[i]["Decimal Places"]
        if pd.isna(v) or str(v).strip() == "": continue
        try: f = float(str(v).strip())
        except (ValueError, TypeError):
            out.append({"row":i+1,"row_index":i,"column":"Decimal Places","issue":f"Invalid value '{v}'","severity":"critical"}); continue
        if f != int(f): out.append({"row":i+1,"row_index":i,"column":"Decimal Places","issue":f"'{v}' not integer","severity":"critical"})
    return out
def validate_mic_sequence(df):
    out = []; order = {n:i for i,n in enumerate(MIC_PATTERN)}; last = -1
    for i in range(len(df)):
        mic = str(df.iloc[i]["MIC Name"]).strip()
        if mic in ("","nan","None"): continue
        if mic not in order:
            out.append({"row":i+1,"row_index":i,"column":"MIC Name","issue":f"Unknown MIC '{mic}'","severity":"warning"}); continue
        if order[mic] < last: out.append({"row":i+1,"row_index":i,"column":"MIC Name","issue":f"MIC sequence break: '{mic}'","severity":"warning"})
        last = order[mic]
    return out
# ── Phase 3: numeric tolerance cross-check vs PDF ──
def validate_limits_against_pdf(df, pdf_rows, alignment_map):
    out = []
    for i, j in enumerate(alignment_map or []):
        if j is None or i >= len(df) or j >= len(pdf_rows): continue
        spec = pdf_rows[j].get("spec",""); exp = parse_limits(spec)
        if not exp: continue
        row = df.iloc[i]
        have, want = _magnums(row["Parameter"]), _magnums(spec)
        if have and want and have != want:
            out.append({"row":i+1,"row_index":i,"column":"Parameter","issue":f"Numbers differ from PDF spec '{spec}'","severity":"warning"})
        for col, key in (("Lower Limit","lower"),("Upper Limit","upper")):
            ev, xv = safe_float(row[col]), exp.get(key)
            if ev is not None and xv is not None and abs(ev-xv) > 1e-6:
                out.append({"row":i+1,"row_index":i,"column":col,"issue":f"Excel {ev} ≠ PDF {xv} (spec '{spec}')","severity":"warning"})
        ed = safe_float(row["Decimal Places"])
        if ed is not None and int(ed) != exp["decimals"]:
            out.append({"row":i+1,"row_index":i,"column":"Decimal Places","issue":f"Excel {int(ed)} ≠ PDF {exp['decimals']}","severity":"warning"})
        xu, eu = exp.get("uom") or "", str(row["UOM"]).strip() if pd.notna(row["UOM"]) else ""
        if xu and eu and _norm(eu) != _norm(xu):
            out.append({"row":i+1,"row_index":i,"column":"UOM","issue":f"Excel '{eu}' ≠ PDF '{xu}'","severity":"warning"})
    return out
# ── Phase 3: MIC category cross-check vs PDF section ──
def validate_mic_against_pdf(df, pdf_rows, alignment_map):
    out = []
    for i, j in enumerate(alignment_map or []):
        if j is None or i >= len(df) or j >= len(pdf_rows): continue
        cat = (pdf_rows[j].get("cat") or "").strip(); mic = str(df.iloc[i]["MIC Name"]).strip()
        if cat and mic and cat != mic:
            out.append({"row":i+1,"row_index":i,"column":"MIC Name","issue":f"Excel '{mic}' ≠ PDF section '{cat}'","severity":"warning"})
    return out
def auto_fill_mic_from_pdf(df, pdf_rows, alignment_map):
    fixed = 0
    for i, j in enumerate(alignment_map or []):
        if j is None or i >= len(df) or j >= len(pdf_rows): continue
        cat = (pdf_rows[j].get("cat") or "").strip(); mic = str(df.iloc[i]["MIC Name"]).strip()
        if cat and mic in ("","nan","None"): df.at[df.index[i],"MIC Name"] = cat; fixed += 1
    return df, fixed
def _session_get(key, default=None):
    try: return st.session_state.get(key, default)
    except Exception: return default
def get_all_issues(df, pv, pp, pm, pdf_rows=None, alignment_map=None):
    all_issues = validate_required_cells(df)+validate_metadata_match(df,pv,pp,pm)+validate_decimal_places(df)+validate_mic_sequence(df)
    pdf_rows = pdf_rows if pdf_rows is not None else _session_get("pdf_rows", [])
    alignment_map = alignment_map if alignment_map is not None else _session_get("alignment_map", [])
    if pdf_rows and alignment_map:
        all_issues += validate_limits_against_pdf(df, pdf_rows, alignment_map)
        all_issues += validate_mic_against_pdf(df, pdf_rows, alignment_map)
    return sorted(all_issues, key=lambda x: {"critical":0,"warning":1}.get(x.get("severity",""),2))

# ════ 9. ALIGNMENT ENGINE (+ Phase 3) ════
@st.cache_resource(show_spinner=False)
def _embed_model():
    from sentence_transformers import SentenceTransformer
    local = os.path.join(os.path.dirname(os.path.abspath(__file__)), "models", "all-MiniLM-L6-v2")
    return SentenceTransformer(local if os.path.isdir(local) else "all-MiniLM-L6-v2")
@st.cache_data(show_spinner=False)
def _encode_texts(texts, model):
    return model.encode(texts, normalize_embeddings=True, batch_size=64, show_progress_bar=False)
_ENUM_PREFIX_RE = re.compile(r"^\s*[\(\[]?\s*\d{1,3}(?:\.\d{1,2})?\s*[\)\]]?\s*\.?\s*")
def _match_norm(text):
    if text is None: return ""
    s = " ".join(_fix_word_token(t) for t in str(text).split())
    s = _ENUM_PREFIX_RE.sub("", s)
    return re.sub(r"[^a-z0-9]","", s.lower())
# ── Phase 3: fuzzy item matching with one-to-one greedy assignment ──
def fallback_align(xl_items, xl_params, pdf_rows, current_match):
    taken = {j for j in current_match if j is not None}
    free = [j for j in range(len(pdf_rows)) if j not in taken]
    cands = []
    for i in range(len(xl_items)):
        if current_match[i] is not None: continue
        ni = _match_norm(xl_items[i])
        if not ni: continue
        for j in free:
            pj = _match_norm(pdf_rows[j].get("item",""))
            if not pj: continue
            str_score = fuzz.token_sort_ratio(ni, pj)/100.0 if RAPIDFUZZ_AVAILABLE else difflib.SequenceMatcher(None, ni, pj).ratio()
            num_score = 0.0
            l1, l2 = parse_limits(xl_params[i]), parse_limits(pdf_rows[j].get("spec",""))
            if l1 and l2 and l1.get("lower") is not None and l2.get("lower") is not None:
                if abs(l1["lower"]-l2["lower"]) < 0.5 and abs((l1.get("upper") or 0)-(l2.get("upper") or 0)) < 0.5: num_score = 0.8
            comb = str_score*0.6 + num_score*0.4
            if comb > 0.6: cands.append((comb, i, j))
    cands.sort(key=lambda t: -t[0])
    ui, uj = set(), set()
    for score, i, j in cands:
        if i in ui or j in uj: continue
        current_match[i] = j; ui.add(i); uj.add(j)
    return current_match
def align_embed(xl_items, pd_items, threshold=0.35):
    if not xl_items or not pd_items: return [None]*len(xl_items)
    xe = pe = None
    try:
        model = _embed_model()
        xe = _encode_texts([str(x) for x in xl_items], model)
        pe = _encode_texts([str(x) for x in pd_items], model)
        if FAISS_AVAILABLE:
            pef = np.asarray(pe).astype(np.float32)
            idx = faiss.IndexFlatIP(pef.shape[1]); idx.add(pef)
            sim = idx.search(np.asarray(xe).astype(np.float32), len(pd_items))[0]
        else: sim = np.asarray(xe) @ np.asarray(pe).T
    except Exception:
        a = [_norm(x) for x in xl_items]; b = [_norm(x) for x in pd_items]
        if RAPIDFUZZ_AVAILABLE:
            from rapidfuzz import process
            match = []
            for it in a:
                r = process.extractOne(it, b, score_cutoff=threshold*100)
                match.append(b.index(r[0]) if r else None)
            return match
        sm = difflib.SequenceMatcher(a=a, b=b, autojunk=False)
        match = [None]*len(a); pa = pb = 0
        for ai, bi, size in sm.get_matching_blocks():
            ga, gb = list(range(pa,ai)), list(range(pb,bi))
            for k, ra in enumerate(ga): match[ra] = gb[k] if k < len(gb) else None
            for k in range(size): match[ai+k] = bi+k
            pa, pb = ai+size, bi+size
        return match
    n, m = len(xl_items), len(pd_items)
    try:
        xn = [_norm(x) for x in xl_items]; pn = [_norm(p) for p in pd_items]
        sm = difflib.SequenceMatcher(autojunk=False)
        for i in range(n):
            if not xn[i]: continue
            sm.set_seq2(xn[i])
            for j in range(m):
                if not pn[j] or sim[i][j] >= 0.72: continue
                sm.set_seq1(pn[j])
                if sm.real_quick_ratio() < 0.72: continue
                cr = sm.ratio()
                if cr > sim[i][j]: sim[i][j] = cr
    except Exception: pass
    dp = [[0.0]*(m+1) for _ in range(n+1)]; bt = [[0]*(m+1) for _ in range(n+1)]
    for i in range(1, n+1):
        row, prev, brow = dp[i], dp[i-1], bt[i]
        for j in range(1, m+1):
            diag = prev[j-1]+float(sim[i-1][j-1]); up, left = prev[j], row[j-1]
            if diag >= up and diag >= left: row[j], brow[j] = diag, 0
            elif up >= left: row[j], brow[j] = up, 1
            else: row[j], brow[j] = left, 2
    match = [None]*n; i, j = n, m
    while i > 0 and j > 0:
        b = bt[i][j]
        if b == 0:
            if float(sim[i-1][j-1]) >= threshold: match[i-1] = j-1
            i -= 1; j -= 1
        elif b == 1: i -= 1
        else: j -= 1
    def same_item(p, q):
        if _norm(xl_items[p]) == _norm(xl_items[q]): return True
        try: return float(np.dot(xe[p], xe[q])) >= 0.75
        except Exception: return False
    for p in range(n):
        if match[p] is None:
            if p > 0 and match[p-1] is not None and same_item(p, p-1): match[p] = match[p-1]
            elif p+1 < n and match[p+1] is not None and same_item(p, p+1): match[p] = match[p+1]
    itm = {}
    for p in range(n):
        if match[p] is not None: itm.setdefault(_norm(xl_items[p]), match[p])
    for p in range(n):
        if match[p] is None:
            k = _norm(xl_items[p])
            if k and k in itm: match[p] = itm[k]
    return match
def match_dims_by_value(xl_params, pdf_rows, m, xl_items=None):
    def _floats(x): return {float(nn) for nn in re.findall(r"\d+(?:\.\d+)?", str(x))}
    pdf_sets = [_floats(r.get("spec","")) for r in pdf_rows]
    item_norm = [_norm(x) for x in xl_items] if xl_items is not None else None
    freq = {}
    if item_norm:
        for nn in item_norm: freq[nn] = freq.get(nn,0)+1
    out = list(m)
    for i, p in enumerate(xl_params):
        if parse_spec(p) is None: continue
        want = _floats(p)
        if len(want) < 2: continue
        j = out[i]
        if j is not None and want <= pdf_sets[j]: continue
        cand = next((k for k,s in enumerate(pdf_sets) if want <= s), None)
        if cand is not None: out[i] = cand; continue
        nominal = max(want)
        if j is not None and nominal in pdf_sets[j]: continue
        if item_norm is not None and j is not None:
            inorm, pit = item_norm[i], _norm(pdf_rows[j].get("item",""))
            if len(inorm) >= 6 and freq.get(inorm,0) == 1 and inorm and pit and (inorm in pit or pit in inorm
             or difflib.SequenceMatcher(None, inorm, pit).ratio() >= 0.6): continue
        out[i] = None
    return out

# ════ 10. UI GRID BUILDERS ════
def _build_comparison_grid_13col(work, pdf_item, pdf_spec, pdf_method, pdf_sampling,
 pdf_vendor="", pdf_part="", pdf_model="", manual_alignments=None, pdf_rows=None):
    disp = work.copy()
    pi, ps, pm_, ps_ = list(pdf_item), list(pdf_spec), list(pdf_method), list(pdf_sampling)
    if manual_alignments and pdf_rows:
        for i, j in manual_alignments.items():
            if j is not None and 0 <= j < len(pdf_rows):
                pi[i] = pdf_rows[j].get("item",""); ps[i] = pdf_rows[j].get("spec","")
                pm_[i] = pdf_rows[j].get("method",""); ps_[i] = pdf_rows[j].get("sampling","")
    pdf_data = {"📄 PDF Vendor":[pdf_vendor]*len(work),"📄 PDF Part No":[pdf_part]*len(work),
     "📄 PDF Model":[pdf_model]*len(work),"📄 PDF Serial No":pi,"📄 PDF MIC":[""]*len(work),
     "📄 PDF Item":pi,"📄 PDF Criteria":ps,"📄 PDF Method":pm_,"📄 PDF Tool":[""]*len(work),
     "📄 PDF Sampling":ps_,"📄 PDF Stage":[""]*len(work),"📄 PDF Remarks":[""]*len(work),
     "📄 PDF Issue Date":[""]*len(work)}
    for ecol, pcol in PDF_COLUMN_MAPPING:
        if ecol in disp.columns:
            disp.insert(disp.columns.get_loc(ecol)+1, pcol, pdf_data.get(pcol, [""]*len(work)))
    return disp
def _comparison_column_config():
    return {pcol: st.column_config.TextColumn(help=f"PDF value (read-only)", disabled=True)
     for ecol, pcol in PDF_COLUMN_MAPPING if pcol in _PDF_READONLY_COLS}
def _apply_excel_edits(work, edited):
    if len(edited):
        work.loc[edited.index, SCHEMA] = edited[SCHEMA].to_numpy()
        work = work.astype(object)
        work["Decimal Places"] = work["Decimal Places"].map(lambda v: None if safe_float(v) == 0 else v)
        st.session_state["work"] = work
    return work
def save_and_validate(work, pv, pp, pm):
    try:
        work["Decimal Places"] = work["Decimal Places"].map(lambda v: None if safe_float(v) == 0 else v)
        st.session_state["work"] = work
        issues = get_all_issues(work, pv, pp, pm)
        st.session_state["issues"] = issues
        return True, issues, sum(1 for i in issues if i.get("severity") == "critical")
    except Exception as e:
        st.error(f"Error saving data: {e}"); return False, [], 0
def reorder_by_mic(df):
    order = {n:i for i,n in enumerate(MIC_PATTERN)}
    return df.sort_values(by="MIC Name", key=lambda x: x.map(lambda m: order.get(str(m).strip(), len(order)))).reset_index(drop=True)

# ════ 11. REPORT DASHBOARD ════
def calculate_metrics(df, pdf_item, pdf_spec, pdf_method, pdf_sampling, issues, auto_summary, pv="", pp="", pm=""):
    total = len(df); matched = sum(1 for s in pdf_spec if str(s).strip())
    crit = sum(1 for i in issues if i.get("severity") == "critical")
    warn = sum(1 for i in issues if i.get("severity") == "warning")
    qs = max(0, min(100, 100-crit*5-warn*2))
    scores = []
    for i in range(len(df)):
        conf = {}
        for col in REQUIRED_COLS:
            if col in df.columns:
                val = str(df.iloc[i][col]) if pd.notna(df.iloc[i][col]) else ""
                pv_ = pdf_item[i] if col == "Inspection Item" and i < len(pdf_item) else (pdf_spec[i] if col == "Parameter" and i < len(pdf_spec) else None)
                conf[col] = calculate_cell_confidence(val, col, pv_)
        scores.append(get_row_confidence_summary(conf)["avg"])
    cm = {}
    for ecol, pcol in PDF_COLUMN_MAPPING:
        if ecol in df.columns:
            mt = 0
            for i in range(len(df)):
                ev = str(df.iloc[i][ecol]).strip() if pd.notna(df.iloc[i][ecol]) else ""
                pv_ = {"📄 PDF Vendor":pv,"📄 PDF Part No":pp,"📄 PDF Model":pm,
                 "📄 PDF Item":pdf_item[i] if i < len(pdf_item) else "",
                 "📄 PDF Criteria":pdf_spec[i] if i < len(pdf_spec) else "",
                 "📄 PDF Method":pdf_method[i] if i < len(pdf_method) else "",
                 "📄 PDF Sampling":pdf_sampling[i] if i < len(pdf_sampling) else ""}.get(pcol,"")
                if (pv_ and ev == pv_) or (not pv_ and not ev): mt += 1
            cm[ecol] = {"matched":mt,"total":total,"percentage":(mt/total*100) if total else 0}
    return {"total_rows":total,"matched_rows":matched,"match_rate":(matched/total*100) if total else 0,
     "quality_score":qs,"critical_issues":crit,"warnings":warn,"issues":issues,"column_match":cm,
     "auto_summary":auto_summary,"can_export":crit == 0,
     "avg_confidence":int(sum(scores)/len(scores)) if scores else 0,
     "low_confidence_rows":sum(1 for s in scores if s < 70)}
def show_kpi_cards(m):
    c1,c2,c3,c4 = st.columns(4)
    with c1: st.markdown(f'<div style="background:linear-gradient(135deg,#1e3a8a,#3b82f6);padding:16px;border-radius:12px;color:white;text-align:center;"><div style="font-size:12px;opacity:.8;">📄 Match Rate</div><div style="font-size:28px;font-weight:bold;">{m["match_rate"]:.0f}%</div><div style="font-size:11px;opacity:.8;">{m["matched_rows"]} of {m["total_rows"]} rows</div></div>', unsafe_allow_html=True)
    with c2:
        q = m["quality_score"]; col = "#10b981" if q >= 80 else "#f59e0b" if q >= 60 else "#ef4444"
        st.markdown(f'<div style="background:linear-gradient(135deg,#065f46,#10b981);padding:16px;border-radius:12px;color:white;text-align:center;"><div style="font-size:12px;opacity:.8;">📊 Quality Score</div><div style="font-size:28px;font-weight:bold;">{q}%</div></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div style="background:linear-gradient(135deg,#7f1d1d,#ef4444);padding:16px;border-radius:12px;color:white;text-align:center;"><div style="font-size:12px;opacity:.8;">⚠ Critical</div><div style="font-size:28px;font-weight:bold;">{m["critical_issues"]}</div></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div style="background:linear-gradient(135deg,#3730a3,#6366f1);padding:16px;border-radius:12px;color:white;text-align:center;"><div style="font-size:12px;opacity:.8;">🎯 Avg. Confidence</div><div style="font-size:28px;font-weight:bold;">{m["avg_confidence"]}%</div></div>', unsafe_allow_html=True)
def show_auto_fixes(a):
    st.markdown("### ✅ Auto-Fixes Applied")
    fd = [("Long Text",a.get("long_text",0)),("Issue Date",a.get("issue_date",0)),("Operation No",a.get("operation_no",0)),
     ("Limits",a.get("limits",0)),("MIC Pattern",a.get("mic_pattern",0)),("Decimal Places",a.get("decimal_places",0))]
    total = sum(c for _,c in fd); mx = max((c for _,c in fd), default=1) or 1
    for name, c in fd:
        st.markdown(f'<div style="display:flex;align-items:center;margin:4px 0;"><div style="width:140px;font-size:13px;">{name}</div><div style="flex:1;height:20px;background:#e2e8f0;border-radius:10px;overflow:hidden;margin:0 10px;"><div style="height:100%;width:{c/mx*100}%;background:linear-gradient(90deg,#3b82f6,#10b981);border-radius:10px;display:flex;align-items:center;justify-content:flex-end;padding-right:8px;color:white;font-size:11px;font-weight:bold;">{c}</div></div><div style="width:80px;font-size:12px;color:#10b981;">✅ Applied</div></div>', unsafe_allow_html=True)
    st.markdown(f"**Total Auto-Fixes: {total}**")
def show_column_match(cm):
    st.markdown("### 🔍 Column Match Analysis")
    for name, d in cm.items():
        p, mt, t = d["percentage"], d["matched"], d["total"]
        col, stt = ("#10b981","✅") if p == 100 else ("#f59e0b","⚠️") if p >= 80 else ("#ef4444","❌")
        st.markdown(f'<div style="display:flex;align-items:center;margin:4px 0;"><div style="width:170px;font-size:13px;">{name}</div><div style="flex:1;height:18px;background:#e2e8f0;border-radius:9px;overflow:hidden;margin:0 10px;"><div style="height:100%;width:{p}%;background:{col};border-radius:9px;display:flex;align-items:center;justify-content:flex-end;padding-right:6px;color:white;font-size:10px;font-weight:bold;">{p:.0f}%</div></div><div style="width:80px;font-size:12px;color:{col};">{stt} {mt}/{t}</div></div>', unsafe_allow_html=True)
def show_issues_table(issues):
    st.markdown("### 📋 Issues Detailed Table")
    if not issues: st.success("✅ No issues found! All data looks perfect."); return
    df = pd.DataFrame([{"Severity":"🔴 CRITICAL" if i.get("severity")=="critical" else "🟡 WARNING","Row":i.get("row",""),"Column":i.get("column",""),"Issue":i.get("issue","")} for i in issues])
    st.dataframe(df, use_container_width=True, hide_index=True)
    crit = sum(1 for i in issues if i.get("severity") == "critical")
    warn = sum(1 for i in issues if i.get("severity") == "warning")
    st.markdown(f'<div style="display:flex;gap:20px;margin-top:10px;padding:10px;background:#f8fafc;border-radius:8px;"><div><span style="color:#ef4444;">🔴 CRITICAL:</span> {crit} issues (must fix before export)</div><div><span style="color:#f59e0b;">🟡 WARNING:</span> {warn} issues (review recommended)</div></div>', unsafe_allow_html=True)

# ════ 12. PIPELINE ORCHESTRATOR (5-stage) ════
@dataclass
class StageResult:
    stage_name: str
    success: bool = True
    processing_time: float = 0.0
    items_processed: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    data: Dict[str, Any] = field(default_factory=dict)
@dataclass
class PipelineResults:
    stage1_preprocess: StageResult = field(default_factory=lambda: StageResult("Preprocessing"))
    stage2_detection: StageResult = field(default_factory=lambda: StageResult("Detection"))
    stage3_recognition: StageResult = field(default_factory=lambda: StageResult("Recognition"))
    stage4_postprocess: StageResult = field(default_factory=lambda: StageResult("Postprocessing"))
    stage5_reconciliation: StageResult = field(default_factory=lambda: StageResult("Reconciliation"))
    total_time: float = 0.0
    success: bool = True
    final_dataframe: Optional[pd.DataFrame] = None
    pdf_rows: List[dict] = field(default_factory=list)
    issues: List[dict] = field(default_factory=list)
    excel_original: Optional[pd.DataFrame] = None
    pdf_metadata: Dict[str, str] = field(default_factory=dict)
    alignment_map: List[Optional[int]] = field(default_factory=list)
class PipelineOrchestrator:
    def __init__(self, pdf_bytes, excel_df, fast=False, use_multi_ocr=True):
        self.pdf_bytes = pdf_bytes; self.excel_original = excel_df.copy()
        self.fast = fast; self.use_multi_ocr = use_multi_ocr
        self.results = PipelineResults()
        self.logger = logging.getLogger("Pipeline")
    def run(self, progress_callback=None):
        t0 = time.time()
        try:
            stages = [("Preprocessing PDF…",0.1,self._stage1),("Detecting layout…",0.2,self._stage2),
             ("Running OCR…",0.3,self._stage3),("Normalizing data…",0.8,self._stage4),
             ("Reconciling with Excel…",0.9,self._stage5)]
            for msg, frac, fn in stages:
                if progress_callback: progress_callback(msg, frac)
                res = fn()
                setattr(self.results, {"Preprocessing":"stage1_preprocess","Detection":"stage2_detection",
                 "Recognition":"stage3_recognition","Postprocessing":"stage4_postprocess",
                 "Reconciliation":"stage5_reconciliation"}[res.stage_name], res)
                if not res.success: self.results.success = False; return self.results
            self.results.total_time = time.time()-t0; self.results.success = True
            self.results.final_dataframe = self.results.stage5_reconciliation.data.get('work_df')
            self.results.pdf_rows = self.results.stage3_recognition.data.get('pdf_rows', [])
            self.results.issues = self.results.stage5_reconciliation.data.get('issues', [])
            self.results.excel_original = self.excel_original
            self.results.pdf_metadata = self.results.stage5_reconciliation.data.get('pdf_metadata', {})
            self.results.alignment_map = self.results.stage5_reconciliation.data.get('alignment_map', [])
            return self.results
        except Exception as e:
            self.logger.error(f"Pipeline failed: {e}", exc_info=True)
            self.results.success = False; return self.results
    def _stage1(self):
        st_ = StageResult("Preprocessing"); t = time.time()
        try:
            pages = pdf_num_pages(self.pdf_bytes); wpp = _digital_words_per_page(self.pdf_bytes)
            st_.data = {'pages':pages,'words_per_page':wpp,'is_scanned':wpp < 25}
            st_.items_processed = pages
        except Exception as e: st_.success = False; st_.errors.append(str(e))
        st_.processing_time = time.time()-t; return st_
    def _stage2(self):
        st_ = StageResult("Detection"); t = time.time()
        st_.data = {'detected_tables':0}; st_.processing_time = time.time()-t; return st_
    def _stage3(self):
        st_ = StageResult("Recognition"); t = time.time()
        try:
            rows, src = extract_pdf_rows(self.pdf_bytes, fast=self.fast)
            st_.data = {'pdf_rows':rows,'source':src,'row_count':len(rows)}; st_.items_processed = len(rows)
        except Exception as e: st_.success = False; st_.errors.append(str(e))
        st_.processing_time = time.time()-t; return st_
    def _stage4(self):
        st_ = StageResult("Postprocessing"); t = time.time()
        rows = self.results.stage3_recognition.data.get('pdf_rows', [])
        st_.data = {'normalized_rows':len(rows)}; st_.items_processed = len(rows)
        st_.processing_time = time.time()-t; return st_
    def _stage5(self):
        st_ = StageResult("Reconciliation"); t = time.time()
        try:
            pdf_rows = self.results.stage3_recognition.data.get('pdf_rows', [])
            work = self.excel_original.copy()
            m = align_embed(list(work["Inspection Item"].astype(str)), [r["item"] for r in pdf_rows])
            m = match_dims_by_value(list(work["Parameter"]), pdf_rows, m, xl_items=list(work["Inspection Item"].astype(str)))
            m = fallback_align(list(work["Inspection Item"].astype(str)), list(work["Parameter"].astype(str)), pdf_rows, m)
            for i in range(1, len(work)):
                op = str(work.iloc[i]["Operation number"]).strip() if pd.notna(work.iloc[i]["Operation number"]) else ""
                if op and re.match(r'^[\(]?[a-zA-Z]\)?', op):
                    cur = work.iloc[i]["Inspection Item"]
                    if pd.isna(cur) or str(cur).strip() == "":
                        prev = work.iloc[i-1]["Inspection Item"]
                        if pd.notna(prev) and str(prev).strip() != "": work.at[work.index[i],"Inspection Item"] = prev
            work, auto = run_automation_pipeline(work)
            work, mic_filled = auto_fill_mic_from_pdf(work, pdf_rows, m)   # Phase 3
            if mic_filled: auto["mic_from_pdf"] = mic_filled
            meta = extract_metadata_from_pdf(self.pdf_bytes)
            pv, pp, pm = meta.get("vendor_code",""), meta.get("part_number",""), meta.get("model_no","")
            issues = get_all_issues(work, pv, pp, pm, pdf_rows, m)        # Phase 3
            st_.data = {'work_df':work,'auto_summary':auto,'issues':issues,'pdf_metadata':meta,
             'pdf_item':[pdf_rows[j]["item"] if j is not None else "" for j in m],
             'pdf_spec':[pdf_rows[j]["spec"] if j is not None else "" for j in m],
             'pdf_method':[pdf_rows[j]["method"] if j is not None else "" for j in m],
             'pdf_sampling':[pdf_rows[j].get("sampling","") if j is not None else "" for j in m],
             'alignment_map':m,'pdf_rows':pdf_rows}
            st_.items_processed = len(work)
        except Exception as e: st_.success = False; st_.errors.append(str(e))
        st_.processing_time = time.time()-t; return st_

# ════ 12b. BATCH PAIRING (new, for Problem 1) ════
##added this in the 1st commit
# Matches uploaded PDFs to uploaded Excels by (vendor_code, part_number).
# vendor_code/part_number for a PDF come from extract_metadata_from_pdf()
# (already defined above); for an Excel, straight from its "VENDOR CODE"
# and "Part number" columns (already part of SCHEMA).
#
# Design decision (confirmed with user): EXACT match only. Any PDF or
# Excel that can't be exactly matched is left unpaired and surfaced to
# the user — never force-paired via fuzzy matching, to avoid silently
# reconciling the wrong PDF against the wrong Excel.

# def _pairing_key_from_pdf(pdf_bytes):
#     """Extract a normalized (vendor_code, part_number) key from a PDF's bytes."""
#     meta = extract_metadata_from_pdf(pdf_bytes)
#     vc, pn = meta.get("vendor_code"), meta.get("part_number")
#     if not vc or not pn: return None
#     return (_norm(vc), _norm(pn))

# def _pairing_key_from_excel(excel_bytes):
#     """Extract a normalized (vendor_code, part_number) key from an Excel's bytes."""
#     try:
#         df = pd.read_excel(io.BytesIO(excel_bytes), dtype=object)
#     except Exception:
#         return None
#     col_map = {_norm(c): c for c in df.columns}
#     vc_col, pn_col = col_map.get(_norm("VENDOR CODE")), col_map.get(_norm("Part number"))
#     if not vc_col or not pn_col: return None
#     vc_series, pn_series = df[vc_col].dropna(), df[pn_col].dropna()
#     if vc_series.empty or pn_series.empty: return None
#     vc, pn = str(vc_series.iloc[0]).strip(), str(pn_series.iloc[0]).strip()
#     if not vc or not pn: return None
#     return (_norm(vc), _norm(pn))

# def pair_files(pdf_files, excel_files):
#     """
#     pdf_files:   list of (filename, raw_bytes) tuples for uploaded PDFs
#     excel_files: list of (filename, raw_bytes) tuples for uploaded Excels
#     Returns {"pairs": [(pdf_filename, excel_filename), ...],
#              "unmatched_pdfs": [...], "unmatched_excels": [...]}
#     """
#     pdf_keys, excel_keys = {}, {}
#     for fname, b in pdf_files:
#         k = _pairing_key_from_pdf(b)
#         if k is not None: pdf_keys[k] = fname
#     for fname, b in excel_files:
#         k = _pairing_key_from_excel(b)
#         if k is not None: excel_keys[k] = fname
#     pairs = [(pdf_fname, excel_keys[k]) for k, pdf_fname in pdf_keys.items() if k in excel_keys]
#     paired_pdf_names = {p for p, e in pairs}
#     paired_excel_names = {e for p, e in pairs}
#     unmatched_pdfs = [fname for fname, b in pdf_files if fname not in paired_pdf_names]
#     unmatched_excels = [fname for fname, b in excel_files if fname not in paired_excel_names]
#     return {"pairs": pairs, "unmatched_pdfs": unmatched_pdfs, "unmatched_excels": unmatched_excels}

def pair_files_serial(pdf_files, excel_files):
    """
    Simple serial pairing: first PDF matches first Excel, second with second, etc.
    pdf_files:   list of (filename, raw_bytes) tuples for uploaded PDFs
    excel_files: list of (filename, raw_bytes) tuples for uploaded Excels
    Returns {"pairs": [(pdf_filename, excel_filename), ...],
             "unmatched_pdfs": [...], "unmatched_excels": [...]}
    """
    pairs = []
    num_pairs = min(len(pdf_files), len(excel_files))
    
    for i in range(num_pairs):
        pairs.append((pdf_files[i][0], excel_files[i][0]))
    
    unmatched_pdfs = [fname for fname, _ in pdf_files[num_pairs:]]
    unmatched_excels = [fname for fname, _ in excel_files[num_pairs:]]
    
    return {"pairs": pairs, "unmatched_pdfs": unmatched_pdfs, "unmatched_excels": unmatched_excels}


# ════ 13. MAIN APP ════

for k, d in [("work",None),("pdf_item",[]),("pdf_spec",[]),("pdf_method",[]),("pdf_sampling",[]),
 ("pdf_vendor",""),("pdf_part",""),("pdf_model",""),("auto_summary",{}),("issues",[]),("_key",None),
 ("_pending",None),("proc_pdf",None),("pdf_src",""),("show_confidence",True),("manual_alignments",{}),
 ("pdf_rows",[]),("alignment_map",[]),("page",1),("rows_per_page",50)]:
    if k not in st.session_state: st.session_state[k] = d


# u1, u2 = st.columns(2)
# pdf_file = u1.file_uploader("📄 PDF — source of truth (SMIR)", type=["pdf"])
# excel_file = u2.file_uploader("📊 Excel — to correct (GSIS-P)", type=["xlsx","xls"])
# if not (pdf_file and excel_file):
#     st.info("⬆️ Upload both a **PDF** and an **Excel** to begin.")
#     st.stop()

#instead of above 6 lines of code, i have changed the code to below
u1, u2 = st.columns(2)
pdf_files_uploaded = u1.file_uploader("📄 PDFs — source of truth (SMIR)", type=["pdf"],
                                       accept_multiple_files=True)
excel_files_uploaded = u2.file_uploader("📊 Excels — to correct (GSIS-P)", type=["xlsx","xls"],
                                         accept_multiple_files=True)
if not (pdf_files_uploaded and excel_files_uploaded):
    st.info("⬆️ Upload one or more **PDFs** and **Excels** to begin (up to 10 each).")
    st.stop()

# # ── Batch pairing preview (Problem 1, step 1) ──
# # Match PDFs to Excels by (Vendor Code, Part Number). Exact match only —
# # unmatched files are surfaced, never force-paired. (pair_files() defined
# # above, in section 12b.)
# _pdf_kv = [(f.name, f.getvalue()) for f in pdf_files_uploaded]
# _excel_kv = [(f.name, f.getvalue()) for f in excel_files_uploaded]
# _pairing = pair_files(_pdf_kv, _excel_kv)

# with st.container(border=True):
#     st.markdown(f"**🔗 Auto-paired: {len(_pairing['pairs'])} pair(s)**")
#     for pdf_name, excel_name in _pairing["pairs"]:
#         st.markdown(f"- 📄 `{pdf_name}` ↔ 📊 `{excel_name}`")
#     if _pairing["unmatched_pdfs"]:
#         st.warning(f"Unmatched PDFs (no Vendor Code + Part Number match found): "
#                    f"{', '.join(_pairing['unmatched_pdfs'])}")
#     if _pairing["unmatched_excels"]:
#         st.warning(f"Unmatched Excels (no Vendor Code + Part Number match found): "
#                    f"{', '.join(_pairing['unmatched_excels'])}")

# if not _pairing["pairs"]:
#     st.error("No pairs could be matched. Reconciliation cannot continue.")
#     st.stop()

# ── Batch pairing preview (serial order) ──
# Serial matching: first PDF ↔ first Excel, second ↔ second, etc.
_pdf_kv = [(f.name, f.getvalue()) for f in pdf_files_uploaded]
_excel_kv = [(f.name, f.getvalue()) for f in excel_files_uploaded]
_pairing = pair_files_serial(_pdf_kv, _excel_kv)

with st.container(border=True):
    st.markdown(f"**🔗 Serial paired: {len(_pairing['pairs'])} pair(s)**")
    for idx, (pdf_name, excel_name) in enumerate(_pairing["pairs"], 1):
        st.markdown(f"- Pair {idx}: 📄 `{pdf_name}` ↔ 📊 `{excel_name}`")
    
    if _pairing["unmatched_pdfs"]:
        st.warning(f"Unmatched PDFs (no corresponding Excel): {', '.join(_pairing['unmatched_pdfs'])}")
    if _pairing["unmatched_excels"]:
        st.warning(f"Unmatched Excels (no corresponding PDF): {', '.join(_pairing['unmatched_excels'])}")

if not _pairing["pairs"]:
    st.error("No pairs could be formed. Please upload at least one PDF and one Excel.")
    st.stop()

# ── Batch reconciliation (Problem 1, step 2) ──
# Resets the batch if the set of matched pairs has changed since last run
# (e.g. user uploaded a different set of files).
_pair_names_now = _pairing["pairs"]
if st.session_state.get("batch_pair_names") != _pair_names_now:
    st.session_state["batch"] = {}
    st.session_state["batch_pair_names"] = _pair_names_now

# if not st.session_state["batch"]:
#     if st.button(f"🚀 Start Reconciliation ({len(_pair_names_now)} pair(s))",
#                  type="primary", use_container_width=True):
#         batch_results = {}
#         with st.status(f"Reconciling {len(_pair_names_now)} pair(s)…", expanded=True) as _status:
#             pb = st.progress(0, text="Starting…")
#             for idx, (pdf_name, excel_name) in enumerate(_pair_names_now):
#                 pb.progress(idx / len(_pair_names_now), text=f"{pdf_name} ↔ {excel_name}…")
#                 pdf_file_i = next(f for f in pdf_files_uploaded if f.name == pdf_name)
#                 excel_file_i = next(f for f in excel_files_uploaded if f.name == excel_name)
#                 pdf_bytes_i = pdf_file_i.getvalue(); excel_bytes_i = excel_file_i.getvalue()
#                 try:
#                     df_raw_i = pd.read_excel(io.BytesIO(excel_bytes_i), dtype=object)
#                 except Exception as e:
#                     st.error(f"Could not read {excel_name}: {e}"); continue
#                 if len(df_raw_i.columns) == len(SCHEMA):
#                     df_raw_i.columns = SCHEMA
#                 else:
#                     n2a = {}
#                     for c in df_raw_i.columns: n2a.setdefault(_norm(c), c)
#                     for c in SCHEMA:
#                         a = n2a.get(_norm(c))
#                         if a is None: df_raw_i[c] = None
#                         elif a != c: df_raw_i = df_raw_i.rename(columns={a: c})
#                     df_raw_i = df_raw_i[SCHEMA]
#                 for _c in df_raw_i.columns: df_raw_i[_c] = df_raw_i[_c].map(strip_residuals)
#                 # Stage A default: reconcile all pages, "Accurate" quality, no multi-OCR.
#                 # (Per-pair page-range / quality controls can be added in a later stage.)
#                 proc_i = subset_pdf(pdf_bytes_i, "")
#                 res = PipelineOrchestrator(proc_i, df_raw_i, fast=False,
#                                             use_multi_ocr=False).run(lambda m, f: None)
#                 if not res.success:
#                     st.error(f"Reconciliation failed for {pdf_name} ↔ {excel_name}."); continue
#                 meta = res.pdf_metadata
#                 batch_results[idx] = {
#                     "pdf_name": pdf_name, "excel_name": excel_name,
#                     "pdf_bytes": pdf_bytes_i, "excel_bytes": excel_bytes_i,
#                     "df_raw": df_raw_i, "proc_pdf": proc_i,
#                     "work": res.final_dataframe,
#                     "auto_summary": res.stage5_reconciliation.data.get('auto_summary', {}),
#                     "issues": res.issues,
#                     "pdf_vendor": meta.get('vendor_code', ''),
#                     "pdf_part": meta.get('part_number', ''),
#                     "pdf_model": meta.get('model_no', ''),
#                     "pdf_rows": res.pdf_rows,
#                     "alignment_map": res.alignment_map,
#                     "pdf_item": res.stage5_reconciliation.data.get('pdf_item', []),
#                     "pdf_spec": res.stage5_reconciliation.data.get('pdf_spec', []),
#                     "pdf_method": res.stage5_reconciliation.data.get('pdf_method', []),
#                     "pdf_sampling": res.stage5_reconciliation.data.get('pdf_sampling', []),
#                     "pdf_src": res.stage3_recognition.data.get('source', ''),
#                     "manual_alignments": {},
#                 }
#             pb.empty()
#             _status.update(label=f"✅ {len(batch_results)}/{len(_pair_names_now)} pair(s) reconciled",
#                             state="complete", expanded=False)
#         st.session_state["batch"] = batch_results
#         st.rerun()
#     st.stop()

if not st.session_state["batch"]:
    if st.button(f"🚀 Start Reconciliation ({len(_pair_names_now)} pair(s))",
                 type="primary", use_container_width=True):
        batch_results = {}
        # Generate a unique batch ID for this entire reconciliation session
        batch_id = generate_unique_id()
        
        with st.status(f"Reconciling {len(_pair_names_now)} pair(s)… (Batch: {batch_id})", expanded=True) as _status:
            pb = st.progress(0, text="Starting…")
            for idx, (pdf_name, excel_name) in enumerate(_pair_names_now):
                # Generate a unique pair ID for each pair within the batch
                pair_id = f"{batch_id}-P{idx+1:02d}"
                pb.progress(idx / len(_pair_names_now), text=f"{pdf_name} ↔ {excel_name} (ID: {pair_id})…")
                
                pdf_file_i = next(f for f in pdf_files_uploaded if f.name == pdf_name)
                excel_file_i = next(f for f in excel_files_uploaded if f.name == excel_name)
                pdf_bytes_i = pdf_file_i.getvalue(); excel_bytes_i = excel_file_i.getvalue()
                try:
                    df_raw_i = pd.read_excel(io.BytesIO(excel_bytes_i), dtype=object)
                except Exception as e:
                    st.error(f"Could not read {excel_name}: {e}"); continue
                if len(df_raw_i.columns) == len(SCHEMA):
                    df_raw_i.columns = SCHEMA
                else:
                    n2a = {}
                    for c in df_raw_i.columns: n2a.setdefault(_norm(c), c)
                    for c in SCHEMA:
                        a = n2a.get(_norm(c))
                        if a is None: df_raw_i[c] = None
                        elif a != c: df_raw_i = df_raw_i.rename(columns={a: c})
                    df_raw_i = df_raw_i[SCHEMA]
                for _c in df_raw_i.columns: df_raw_i[_c] = df_raw_i[_c].map(strip_residuals)
                proc_i = subset_pdf(pdf_bytes_i, "")
                res = PipelineOrchestrator(proc_i, df_raw_i, fast=False,
                                            use_multi_ocr=False).run(lambda m, f: None)
                if not res.success:
                    st.error(f"Reconciliation failed for {pdf_name} ↔ {excel_name}."); continue
                meta = res.pdf_metadata
                batch_results[idx] = {
                    "pair_id": pair_id,  # NEW: Add unique pair ID
                    "batch_id": batch_id,  # NEW: Add batch ID
                    "pdf_name": pdf_name, "excel_name": excel_name,
                    "pdf_bytes": pdf_bytes_i, "excel_bytes": excel_bytes_i,
                    "df_raw": df_raw_i, "proc_pdf": proc_i,
                    "work": res.final_dataframe,
                    "auto_summary": res.stage5_reconciliation.data.get('auto_summary', {}),
                    "issues": res.issues,
                    "pdf_vendor": meta.get('vendor_code', ''),
                    "pdf_part": meta.get('part_number', ''),
                    "pdf_model": meta.get('model_no', ''),
                    "pdf_rows": res.pdf_rows,
                    "alignment_map": res.alignment_map,
                    "pdf_item": res.stage5_reconciliation.data.get('pdf_item', []),
                    "pdf_spec": res.stage5_reconciliation.data.get('pdf_spec', []),
                    "pdf_method": res.stage5_reconciliation.data.get('pdf_method', []),
                    "pdf_sampling": res.stage5_reconciliation.data.get('pdf_sampling', []),
                    "pdf_src": res.stage3_recognition.data.get('source', ''),
                    "manual_alignments": {},
                }
            pb.empty()
            _status.update(label=f"✅ {len(batch_results)}/{len(_pair_names_now)} pair(s) reconciled (Batch: {batch_id})",
                            state="complete", expanded=False)
        st.session_state["batch"] = batch_results
        st.rerun()
    st.stop()

# ── Per-pair result tabs (Problem 1, step 2) ──
# Each matched pair gets its own outer tab; the same 5 inner tabs from the
# original single-pair app appear inside each one, driven by that pair's
# own stored results (never the shared flat keys the original app used).
_pair_ids = sorted(st.session_state["batch"].keys())
# _outer_labels = [f"✅ {st.session_state['batch'][pid]['pdf_name']}" for pid in _pair_ids]

_outer_labels = [
    f"✅ {st.session_state['batch'][pid]['pdf_name']} (ID: {st.session_state['batch'][pid].get('pair_id', 'N/A')})" 
    for pid in _pair_ids
]
_outer_tabs = st.tabs(_outer_labels)
for _oi, _pid in enumerate(_pair_ids):
    with _outer_tabs[_oi]:
        _pd = st.session_state["batch"][_pid]
        
        # Display the unique ID prominently
        st.markdown(f"""
        <div style="background: #f0f4ff; padding: 10px 16px; border-radius: 8px; border-left: 4px solid #3b82f6; margin-bottom: 16px;">
            <span style="font-size: 13px; color: #64748b;">🆔 Pair ID:</span>
            <span style="font-size: 14px; font-weight: 600; color: #1e3a8a; font-family: monospace;">
                {_pd.get('pair_id', 'N/A')}
            </span>
            <span style="margin-left: 20px; font-size: 13px; color: #64748b;">📦 Batch:</span>
            <span style="font-size: 14px; font-weight: 500; color: #1e3a8a; font-family: monospace;">
                {_pd.get('batch_id', 'N/A')}
            </span>
        </div>
        """, unsafe_allow_html=True)
        
        pdf_bytes = _pd["pdf_bytes"]; excel_bytes = _pd["excel_bytes"]
        df_raw = _pd["df_raw"]; df_orig = df_raw
        alignment_map = _pd.get("alignment_map", [])
        work = _pd["work"]
        # ... rest of the code continues as before
        if any(str(work[c].dtype) != "object" for c in work.columns):
            work = work.astype(object); st.session_state["batch"][_pid]["work"] = work
        work["Decimal Places"] = work["Decimal Places"].map(lambda v: None if safe_float(v) == 0 else v)
        df_orig = df_raw
        pdf_item = _pd["pdf_item"]; pdf_spec = _pd["pdf_spec"]
        pdf_method = _pd["pdf_method"]; pdf_sampling = _pd["pdf_sampling"]
        pdf_vendor = _pd.get("pdf_vendor",""); pdf_part = _pd.get("pdf_part","")
        pdf_model = _pd.get("pdf_model",""); auto_summary = _pd.get("auto_summary",{})
        pdf_rows = _pd.get("pdf_rows",[]); manual_alignments = _pd.get("manual_alignments",{})
        issues = get_all_issues(work, pdf_vendor, pdf_part, pdf_model)
        st.session_state["batch"][_pid]["issues"] = issues
        metrics = calculate_metrics(work, pdf_item, pdf_spec, pdf_method, pdf_sampling, issues, auto_summary, pdf_vendor, pdf_part, pdf_model)
        tab_report, tab_review, tab_edit, tab_pdf, tab_export = st.tabs(["📊 Report Dashboard","🔍 Review & Fix","✏️ Edit Excel","📄 PDF (source of truth)","📤 Export"])
        with tab_report:
            st.markdown("## 📊 Reconciliation Report Dashboard")
            show_kpi_cards(metrics); st.divider()
            c1, c2 = st.columns(2)
            with c1: show_auto_fixes(auto_summary)
            with c2: show_column_match(metrics["column_match"])
            st.divider(); show_issues_table(issues)
        # with tab_review:
        #     st.markdown("### 🔍 Review & Fix")
        #     ct, cf, cs = st.columns([1,2,3])
        #     with ct: st.session_state["show_confidence"] = st.checkbox("📊 Show Confidence", value=st.session_state.get("show_confidence",True))
        #     with cf: flag = st.radio("Show", ["All rows","Rows with issues","Rows with PDF match","Rows missing PDF match"], horizontal=True, index=0, label_visibility="collapsed")
        #     with cs: search = st.text_input("search", placeholder="🔎 Search…", label_visibility="collapsed", key=f"review_search_{_pid}")
        with tab_review:
    st.markdown("### 🔍 Review & Fix")
    ct, cf, cs = st.columns([1,2,3])
    with ct: 
        st.session_state["show_confidence"] = st.checkbox(
            "📊 Show Confidence", 
            value=st.session_state.get("show_confidence", True),
            key=f"show_conf_{_pid}"  # Add this unique key
        )
    with cf: 
        flag = st.radio(
            "Show", 
            ["All rows","Rows with issues","Rows with PDF match","Rows missing PDF match"], 
            horizontal=True, 
            index=0, 
            label_visibility="collapsed",
            key=f"review_radio_{_pid}"  # Also add unique key to radio
        )
    with cs: 
        search = st.text_input(
            "search", 
            placeholder="🔎 Search…", 
            label_visibility="collapsed", 
            key=f"review_search_{_pid}"
        )
            mask = np.ones(len(work), dtype=bool)
            if flag == "Rows with issues":
                ir = {i["row_index"] for i in issues}; mask = np.array([i in ir for i in range(len(work))])
            elif flag == "Rows with PDF match": mask &= np.array([bool(str(s).strip()) for s in pdf_spec])
            elif flag == "Rows missing PDF match": mask &= np.array([not str(s).strip() for s in pdf_spec])
            if search.strip():
                s = search.strip().lower()
                mask &= (work["Inspection Item"].astype(str).str.lower().str.contains(s,regex=False) | work["Parameter"].astype(str).str.lower().str.contains(s,regex=False)).to_numpy()
            disp = _build_comparison_grid_13col(work, pdf_item, pdf_spec, pdf_method, pdf_sampling, pdf_vendor, pdf_part, pdf_model, manual_alignments, pdf_rows)
            view_full = disp[mask]; total_rows = len(view_full)
            rpp_opts = [25,50,100,200]
            # 
            rpp = st.selectbox("Rows per page", options=rpp_opts, index=rpp_opts.index(_pd.get("rows_per_page",50)), key=f"rpp_sel_{_pid}")
            tp = max(1, (total_rows+rpp-1)//rpp)
            c1,c2,c3 = st.columns([1,2,1])
            with c1:
                if st.button("◀ Previous", disabled=_pd.get("page",1) <= 1): _pd["page"] = max(1, _pd.get("page",1)-1); st.rerun()
            # with c2: _pd["page"] = st.number_input("Page", 1, tp, min(tp, _pd.get("page",1)), 1, key=f"page_in_{_pid}", label_visibility="collapsed")
            with c2: _pd["page"] = st.number_input("Page", 1, tp, min(tp, _pd.get("page",1)), 1, key=f"page_in_{_pid}", label_visibility="collapsed")
            with c3:
                if st.button("Next ▶", disabled=_pd.get("page",1) >= tp): _pd["page"] = min(tp, _pd.get("page",1)+1); st.rerun()
            si = (_pd.get("page",1)-1)*rpp; view = view_full.iloc[si:min(si+rpp, total_rows)]
            st.caption(f"Rows {si+1}–{min(si+rpp,total_rows)} of {total_rows}")
            if st.session_state["show_confidence"]:
                cd = []
                for i in range(len(view)):
                    ri = view.index[i]; rc = {}
                    for col in REQUIRED_COLS:
                        if col in view.columns:
                            val = str(view.iloc[i][col]) if pd.notna(view.iloc[i][col]) else ""
                            pv_ = pdf_item[ri] if col == "Inspection Item" and ri < len(pdf_item) else (pdf_spec[ri] if col == "Parameter" and ri < len(pdf_spec) else None)
                            rc[col] = calculate_cell_confidence(val, col, pv_)
                    sm = get_row_confidence_summary(rc)
                    cd.append({"Avg":sm["avg"],"Min":sm["min"],"Needs Review":"⚠️" if sm["needs_review"] else "✅"})
                cdf = pd.DataFrame(cd)
                view = view.copy()
                for col in ["Avg","Min","Needs Review"]: view.insert(1, f"📊 {col}", cdf[col])
                cfg = _comparison_column_config()
                cfg["📊 Avg"] = st.column_config.NumberColumn("Avg Conf", format="%d%%")
                cfg["📊 Min"] = st.column_config.NumberColumn("Min Conf", format="%d%%")
                cfg["📊 Needs Review"] = st.column_config.TextColumn("Review")
                disabled = list(_PDF_READONLY_COLS)+["📊 Avg","📊 Min"," Needs Review"]
            else:
                cfg = _comparison_column_config(); disabled = list(_PDF_READONLY_COLS)
            edited = st.data_editor(view, use_container_width=True, height=470, num_rows="fixed", disabled=disabled, column_config=cfg, key=f"review_grid_{_pid}")
            work = _apply_excel_edits(work, edited)
            cb1, cb2 = st.columns(2)
            with cb1:
                if st.button("⭮ Renumber Operation №", use_container_width=True):
                    c = 1
                    for i in range(len(work)):
                        if _row_has_inspection_content(work.iloc[i]): work.at[work.index[i],"Operation number"] = c; c += 1
                    st.session_state["batch"][_pid]["work"] = work; st.rerun()
            with cb2:
                if st.button("🔄 Reorder by MIC", use_container_width=True):
                    work = reorder_by_mic(work); st.session_state["batch"][_pid]["work"] = work; st.rerun()
            unmatched = [i for i, s in enumerate(pdf_spec) if not str(s).strip()]
            if unmatched:
                st.markdown("### 🔄 Manual Alignment Override")
                with st.container(border=True):
                    opts = ["None"]+[f"{j+1}: {pdf_rows[j].get('item','')[:40]}" for j in range(len(pdf_rows))]
                    idxs = [-1]+list(range(len(pdf_rows)))
                    with st.form(f"manual_align_form_{_pid}"):
                        tmp = {}
                        for ri in unmatched:
                            cols = st.columns([1,2,2,3])
                            cols[0].write(f"{ri+1}"); cols[1].write(str(work.iloc[ri]["Inspection Item"])[:40]); cols[2].write(str(work.iloc[ri]["Parameter"])[:40])
                            cur = manual_alignments.get(ri, -1)
                            di = idxs.index(cur) if cur in idxs else 0
                            sel = cols[3].selectbox("Select", options=opts, index=di, key=f"ma_{_pid}_{ri}", label_visibility="collapsed")
                            tmp[ri] = idxs[opts.index(sel)]
                        ca, cc = st.columns(2)
                        if ca.form_submit_button("✅ Apply", use_container_width=True):
                            st.session_state["batch"][_pid]["manual_alignments"] = tmp; st.rerun()
                        if cc.form_submit_button("🗑️ Clear", use_container_width=True):
                            st.session_state["batch"][_pid]["manual_alignments"] = {}; st.rerun()
        with tab_edit:
            st.markdown("### ✏️ Edit Excel")
            b1,b2,b3,b4 = st.columns([1,1,1,2])
            with b1:
                if st.button("💾 Save Changes", type="primary", use_container_width=True):
                    ok, ni, cr = save_and_validate(work, pdf_vendor, pdf_part, pdf_model)
                    if ok: st.success(f"✅ Saved! {cr} critical."); st.rerun()
            with b2:
                if st.button("🔄 Re-apply Automation", use_container_width=True):
                    wu, ns, fc = reapply_automation(work)
                    st.session_state["batch"][_pid]["work"] = wu; st.session_state["batch"][_pid]["auto_summary"] = ns; st.success(f"✅ Fixed {fc} cells."); st.rerun()
            with b3:
                if st.button("⭮ Renumber Operations", use_container_width=True):
                    c = 1
                    for i in range(len(work)):
                        if _row_has_inspection_content(work.iloc[i]): work.at[work.index[i],"Operation number"] = c; c += 1
                    st.session_state["batch"][_pid]["work"] = work; st.rerun()
            with b4: st.caption("💡 Edit any cell, then Save.")
            st.divider()
            c1, c2 = st.columns([1,3])
            ef = c1.radio("Show", ["All rows","Rows with issues"], key=f"edit_filter_{_pid}", label_visibility="collapsed")
            es = c2.text_input("edit_search", placeholder="🔎 Search…", label_visibility="collapsed", key=f"edit_search_{_pid}")
            em = np.ones(len(work), dtype=bool)
            if ef == "Rows with issues":
                ir = {i["row_index"] for i in issues}; em = np.array([i in ir for i in range(len(work))])
            if es.strip():
                s = es.strip().lower()
                em &= (work["Inspection Item"].astype(str).str.lower().str.contains(s,regex=False) | work["Parameter"].astype(str).str.lower().str.contains(s,regex=False)).to_numpy()
            ev = work[em]
            ee = st.data_editor(ev[SCHEMA], use_container_width=True, height=520, num_rows="fixed", key=f"edit_grid_{_pid}")
            if not ee.equals(ev): work = _apply_excel_edits(work, ee)
        with tab_pdf:
            st.markdown("### 📄 PDF (source of truth)")
            c1,c2,c3 = st.columns([1,2,1])
            with c2:
                if st.button("🔄 Re-extract Metadata", use_container_width=True):
                    with st.spinner("Re-extracting…"):
                        m = extract_metadata_from_pdf(st.session_state.get("proc_pdf", pdf_bytes))
                        st.session_state["batch"][_pid]["pdf_vendor"] = m.get("vendor_code","")
                        st.session_state["batch"][_pid]["pdf_part"] = m.get("part_number","")
                        st.session_state["batch"][_pid]["pdf_model"] = m.get("model_no","")
                        st.success("Metadata updated!"); st.rerun()
            pages = pdf_page_pngs(st.session_state.get("proc_pdf", pdf_bytes))
            if pages:
                c1,c2,c3 = st.columns([1,1,2])
                pn = c1.number_input("Page", 1, len(pages), 1, 1)
                fit = c2.selectbox("Fit", ["Width","Custom"], 0)
                zm = c3.slider("Zoom %", 40, 300, 100, 10, disabled=fit == "Width")
                wcss = "100%" if fit == "Width" else f"{zm}%"
                st.markdown(f'<div style="overflow:auto;max-height:80vh;border:1px solid #cbd5e1;border-radius:10px;background:#f8fafc;"><img src="data:image/png;base64,{base64.b64encode(pages[int(pn)-1]).decode()}" style="width:{wcss};display:block;margin:auto"></div>', unsafe_allow_html=True)
            else: st.warning("Could not render the PDF.")
            with st.expander("📋 Extracted rows"):
                st.dataframe(pd.DataFrame({"PDF Item":pdf_item,"PDF Spec":pdf_spec,"PDF Method":pdf_method,"PDF Sampling":pdf_sampling}), use_container_width=True, height=300)
        with tab_export:
            st.markdown("### 📤 Export & Download")
            crit = [i for i in issues if i.get("severity") == "critical"]
            if crit: st.warning(f"{len(crit)} critical issue(s) remain — review before export.")
            _MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            try:
                corrected = build_workbook(excel_bytes, work)
                rebuilt = _to_xlsx(build_rebuilt(pdf_rows,
                 {"vendor_code":pdf_vendor,"part_numbers":[pdf_part],"model_no":pdf_model},
                 work.iloc[0]["Issue date"] if len(work) else None))
                diffcsv = build_diff(pdf_rows, work, alignment_map).to_csv(index=False).encode("utf-8")
                c1,c2,c3 = st.columns(3)
                c1.download_button("⬇ Corrected Excel", corrected, file_name="corrected.xlsx", mime=_MIME, use_container_width=True)
                c2.download_button("⬇ Rebuilt-from-PDF", rebuilt, file_name="rebuilt.xlsx", mime=_MIME, use_container_width=True)
                c3.download_button("⬇ Diff (CSV)", diffcsv, file_name="diff.csv", mime="text/csv", use_container_width=True)
                st.caption("Works from any PC on your LAN — open http://<host-ip>:8502 and click a download button.")
            except Exception as e:
                st.error(f"Export failed: {type(e).__name__}: {e}")
                # CSV fallback so you can ALWAYS get the data
                st.download_button("⬇ Fallback CSV", work[SCHEMA].to_csv(index=False).encode("utf-8"),
                 file_name="corrected.csv", mime="text/csv", use_container_width=True)

                 ##checkkkk